#Importação de bibliotecas

import datetime
import os
import sqlite3
import sys
from datetime import date, datetime
from tkinter import *
from tkinter import IntVar, PhotoImage, StringVar, scrolledtext
from tkinter import Tk as tk
from tkinter import filedialog
from tkinter import filedialog as fd
from tkinter import messagebox, ttk
from tkinter.filedialog import asksaveasfilename
from tkinter.font import Font
import openpyxl
import pandas as pd
import pyperclip
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image, ImageTk
from tkcalendar import Calendar, DateEntry
import json
import configparser
import re
import locale
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')  # Define a formatação local para o Brasil



sys.path.append(
    r"\\192.168.16.4\central de vendas\cadapp\.venv\cadapp_v.01\main_teste1"
)

print(
    os.path.exists(
        r"\\192.168.16.4\central de vendas\cadapp\.venv\cadapp_v.01\main_teste1\teste_loggin_DataBaser.py"
    )
)


import configparser
import pickle
import pickletools
import time
import tkinter


###################
def conectar_banco():
    try:
        # Conectar ao banco de dados
        conn = sqlite3.connect(
            r"\\192.168.16.4\central de vendas\cadapp\atendimento_wpp_teste1.db"
        )
        cursor = conn.cursor()

        # Criar a tabela se ela não existir
        cursor.execute(
            """
        CREATE TABLE IF NOT EXISTS users_teste1(
            Id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT,
            Nome TEXT NOT NULL,
            Email TEXT NOT NULL,
            Usuario TEXT NOT NULL,
            Senha TEXT NOT NULL
        );
        """
        )

        # print('Conectado ao Banco de Dados e tabela criada/verificada com sucesso!')
        return conn, cursor

    except sqlite3.Error as e:
        print(f"Erro ao conectar ao banco de dados: {e}")
        return None, None


# Certifique-se de importar ou definir a função conectar_banco
conn, cursor = conectar_banco()  # Chame a função para obter o cursor

# cores
cod_cor01 = "#398564"  # Verde floresta
cod_cor02 = "#82b84d"  # Verde claro
cod_cor03 = "#1f4f29"  # Verde escuro
cod_cor14 = "#ffffff"  # Branco
cod_cor15 = "#f9f7f3"  # Branco quase creme
cod_cor26 = "#8ea975"  # cor logo background

# Função para obter o caminho absoluto, seja em desenvolvimento ou em executável empacotado
def resource_path(relative_path):
    """Obter o caminho absoluto, usa _MEIPASS quando o programa está empacotado"""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(os.path.dirname(__file__))

    return os.path.join(base_path, relative_path)


# Caminho para o banco de dados (no mesmo diretório do executável)
db_path = r"\\192.168.16.4\central de vendas\cadapp\atendimento_wpp_teste1.db"

# Verificar se o banco de dados já existe antes de conectar
if os.path.exists(db_path):
    print(f"Banco de dados encontrado: {db_path}")
else:
    print(f"Erro: O banco de dados não foi encontrado no caminho {db_path}")


# Conectar ao banco de dados
con = sqlite3.connect(db_path)

# Diretório atual (onde o script está sendo executado)
base_dir_atual = os.path.dirname(os.path.abspath(__file__))
# print(f"Diretório atual: {base_dir_atual}")
global tree
global imagem_string, imagem, l_imagem
global hoje, agora

####### CRIANDO janela_loggin ##############################################################
janela_loggin = Tk()
janela_loggin.title("Teste Loggin 1 Acess Panel")
janela_loggin.geometry("600x300")
janela_loggin.configure(background=cod_cor15)
janela_loggin.resizable(FALSE, FALSE)
janela_loggin.attributes("-alpha", 1)
janela_loggin.iconbitmap(
    default=r"\\192.168.16.4\central de vendas\cadapp\.venv\cadapp_v.01\main_teste1\imgs\logo_loggin_icon.ico"
)

####### FCARREGAR IMAGEM LOGO ##############################################################
logo = Image.open(
    r"\\192.168.16.4\central de vendas\cadapp\.venv\cadapp_v.01\main_teste1\imgs\logo_loggin.png"
)
logo = logo.resize((150, 150))
logo = ImageTk.PhotoImage(logo)

####### FRAMES / WIDGETS ##############################################################
frame_esquerda = Frame(
    janela_loggin, width=200, height=300, bg=cod_cor03, relief="raise"
)
frame_esquerda.pack(side=LEFT)

frame_direita = Frame(
    janela_loggin, width=395, height=300, bg=cod_cor03, relief="raise"
)
frame_direita.pack(side=RIGHT)


####### LABELS E ENTRYS ##############################################################

# LOGO
logo_label = Label(frame_esquerda, image=logo, bg=cod_cor03)
logo_label.place(x=20, y=90)

# USUÁRIO
# label
user_label = Label(
    frame_direita,
    text="Usuário:",
    font=("Century Gothic", 20, "bold"),
    bg=cod_cor03,
    fg=cod_cor15,
)
user_label.place(x=10, y=90)

global user_entry
# entry
user_entry = ttk.Entry(frame_direita, width=30)
user_entry.place(x=135, y=102)

# SENHA
# label
senha_label = Label(
    frame_direita,
    text="Senha:",
    font=("Century Gothic", 20, "bold"),
    bg=cod_cor03,
    fg=cod_cor15,
)
senha_label.place(x=10, y=143)
# entry
senha_entry = ttk.Entry(frame_direita, width=30, show="*")
senha_entry.place(x=135, y=155)
import tkinter as tk

# Variável global para armazenar o usuário logado
global usuario_logado




# Conectar ao banco de dados e inicializar `conn` e `cursor`
conn = sqlite3.connect(
    r"\\192.168.16.4\central de vendas\cadapp\atendimento_wpp_teste1.db"
)
cursor = conn.cursor()

"""# Função para salvar os dados de login (usuário e senha)
def salvar_dados_login(usuario, senha):
    dados_login = {"usuario": usuario, "senha": senha}
    with open("dados_login.pkl", "wb") as arquivo:
        pickle.dump(dados_login, arquivo)

# Função para carregar os dados de login (se já estiverem salvos)
def carregar_dados_login():
    try:
        with open("dados_login.pkl", "rb") as arquivo:
            dados_login = pickle.load(arquivo)
        return dados_login
    except FileNotFoundError:
        return None"""


# teste_loggin_DataBaser
# Função para preencher dados salvos
def preencher_dados_salvos():
    pass  # Marca o checkbox automaticamente


# REGISTRAR
# Função de Registro
def register():
    # Remover botões de login
    login_button.place(x=80000)
    registrar_button.place(x=8000)

    # Inserir itens de cadastro
    nome_label = Label(
        frame_direita,
        text="Nome:",
        font=("Century Gothic", 20, "bold"),
        bg=cod_cor03,
        fg=cod_cor15,
    )
    nome_label.place(x=10, y=5)
    nome_entry = ttk.Entry(frame_direita, width=30)
    nome_entry.place(x=135, y=17)

    email_label = Label(
        frame_direita,
        text="Email:",
        font=("Century Gothic", 20, "bold"),
        bg=cod_cor03,
        fg=cod_cor15,
    )
    email_label.place(x=10, y=45)
    email_entry = ttk.Entry(frame_direita, width=30)
    email_entry.place(x=135, y=57)

    # Função interna para registrar no banco de dados
    def registrar_no_db():
        nome = nome_entry.get()
        email = email_entry.get()
        Usuario = user_entry.get()
        Senha = senha_entry.get()

        # Verificar se todos os campos foram preenchidos
        if nome == "" or email == "" or Usuario == "" or Senha == "":
            messagebox.showerror("Erro", "Preencha TODOS os campos")
        else:
            # Inserir dados no banco de dados
            cursor.execute(
                """
            INSERT INTO users_teste1 (Nome, Email, Usuario, Senha)
            VALUES (?, ?, ?, ?);""",
                (nome, email, Usuario, Senha),
            )

            # Confirmar a transação no banco
            conn.commit()
            messagebox.showinfo("Register Info", "Conta criada com sucesso!")

    # Botão para confirmar o registro
    registrar_func = ttk.Button(
        frame_direita, text="Registrar", width=30, command=registrar_no_db
    )
    registrar_func.place(x=135, y=225)

    # função voltar
    def voltar_loggin():
        # removendo labels
        nome_label.place(x=8000)
        nome_entry.place(x=8000)
        email_label.place(x=8000)
        email_entry.place(x=8000)
        registrar_func.place(x=8000)
        voltar_func_reg.place(x=8000)
        # Trazendo labels login
        login_button.place(x=135, y=225)
        registrar_button.place(x=170)

    # Voltar
    voltar_func_reg = ttk.Button(
        frame_direita, text="Voltar", width=20, command=voltar_loggin
    )
    voltar_func_reg.place(x=170, y=260)


# LOGGIN


# Função de login
def loggin():
    global usuario_logado
    Usuario = user_entry.get()
    Senha = senha_entry.get()

    cursor.execute(
        """
        SELECT Usuario, Senha
        FROM users_teste1
        WHERE (Usuario = ? AND Senha = ?)
    """,
        (Usuario, Senha),
    )

    verificar_login = cursor.fetchone()

    if verificar_login is not None:
        if Usuario in verificar_login and Senha in verificar_login:
            # Atualiza o usuário logado e inicia a próxima janela
            usuario_logado = Usuario
            iniciar_janela()  # Supondo que essa função inicia a próxima janela
        else:
            messagebox.showwarning("Aviso de Login", "Usuário ou senha incorretos!")
    else:
        messagebox.showerror(
            "Aviso de Login", "Acesso negado, usuário ou senha incorretos!"
        )


# Variável para o estado da caixa de seleção "Manter-me conectado"
var_mantenha = IntVar()

# print("Selecionou")


# Função para salvar os dados de login (usuário e senha)
def salvar_dados_login(usuario, senha):
    dados_login = {"usuario": usuario, "senha": senha}
    with open("dados_login.pkl", "wb") as arquivo:
        pickle.dump(dados_login, arquivo)


# Função para carregar os dados de login (se já estiverem salvos)
def carregar_dados_login():
    try:
        with open("dados_login.pkl", "rb") as arquivo:
            dados_login = pickle.load(arquivo)
        return dados_login
    except FileNotFoundError:
        return None


def preencher_dados_salvos():
    # Consulta para verificar se existe algum usuário com 'lembrar_login' ativado
    cursor.execute(
        """
        SELECT Usuario, Senha
        FROM users_teste1
        WHERE lembrar_login = 1
    """
    )

    usuario_lembrado = cursor.fetchone()

    # Se um usuário com 'lembrar_login' ativado for encontrado, preenche os campos de login
    if usuario_lembrado:
        Usuario, Senha = usuario_lembrado
        user_entry.insert(0, Usuario)
        senha_entry.insert(0, Senha)
        var_mantenha.set(1)  # Marca o checkbox automaticamente


# Após verificar login com sucesso

usuario_logado = user_entry.get()  # Nome do usuário que acabou de logar


# Função fechar app


def fechar_app():
    janela_loggin.destroy()
    sys.exit()


def iniciar_janela():
    try:
        # FORMATACAO PARA  A INTERFACE

        import datetime
        import operator
        import os
        import sqlite3
        import sys
        import tkinter
        from datetime import date
        from tkinter import (CENTER, FLAT, LEFT, NW, RIDGE, VERTICAL, Button,
                             Canvas, Entry, Frame, Label, PhotoImage,
                             Radiobutton, Scrollbar, StringVar)
        from tkinter import Tk
        from tkinter import Tk as tk
        from tkinter import Toplevel
        from tkinter import filedialog
        from tkinter import filedialog as fd
        from tkinter import messagebox, ttk
        from tkinter.filedialog import (askopenfile, askopenfilename,
                                        askopenfilenames, asksaveasfilename)
        from tkinter.font import Font

        import openpyxl
        import pandas as pd
        import PIL
        import pyperclip
        from openpyxl.styles import NamedStyle
        from openpyxl.utils.dataframe import dataframe_to_rows
        from PIL import Image, ImageFile, ImageTk
        from tkcalendar import Calendar, DateEntry

        # cores
        cod_cor01 = "#398564"  # Verde floresta
        cod_cor02 = "#82b84d"  # Verde claro
        cod_cor03 = "#1f4f29"  # Verde escuro
        cod_cor04 = "#e6e6e6"  # Cinza claro
        cod_cor05 = "#a6a6a6"  # Cinza médio
        cod_cor06 = "#4d4d4d"  # Cinza escuro
        cod_cor07 = "#ffcc33"  # Amarelo médio
        cod_cor08 = "#ffdd66"  # Amarelo claro
        cod_cor09 = "#ff9933"  # Laranja suave
        cod_cor10 = "#292b2c"  # Preto carvão
        cod_cor11 = "#343a40"  # Cinza antracito
        cod_cor12 = "#595959"  # Cinza neutro
        cod_cor13 = "#b3b3b3"  # Cinza pálido
        cod_cor14 = "#ffffff"  # Branco
        cod_cor15 = "#f9f7f3"  # Branco quase creme
        cod_cor16 = "#7f8c50"  # Verde militar
        cod_cor17 = "#cdcdcd"  # Cinza concreto
        cod_cor18 = "#ffb84d"  # Laranja suave
        cod_cor19 = "#996633"  # Marrom suave
        cod_cor20 = "#e0d5b6"  # Bege claro
        cod_cor21 = "#5c5c5c"  # Cinza aço
        cod_cor22 = "#FFF3CF"  # amarelo limao fraco
        cod_cor23 = "#f2f2f2"  # Cinza muito claro
        cod_cor24 = "#ffcc99"  # Salmão claro
        cod_cor25 = "#c9c9c9"  # Cinza suave
        cod_cor26 = "#8ea975"  # cor logo backgound
        cod_cor27 = "#E6F0DC"  # verde bem claro tipo chantilly
        cod_cor28 = "#003300"  # verde escuro

        # FUNÇÕES de conecção ###########################################################################################################################################
        global db_path
        global resource_path
        # diretório do DB
        janela_loggin.attributes("-alpha", 0)

        # Função para obter o caminho absoluto, seja em desenvolvimento ou em executável empacotado
        # Função para obter o caminho absoluto, seja em desenvolvimento ou em executável empacotado
        def resource_path(relative_path):
            """Obter o caminho absoluto, usa _MEIPASS quando o programa está empacotado"""
            try:
                base_path = sys._MEIPASS
            except AttributeError:
                base_path = os.path.abspath(os.path.dirname(__file__))

            return os.path.join(base_path, relative_path)

        # Caminho para o banco de dados (no mesmo diretório do executável)
        db_path = r"\\192.168.16.4\central de vendas\cadapp\atendimento_wpp_teste1.db"

        # Verificar se o banco de dados já existe antes de conectar
        if os.path.exists(db_path):
            print(f"Banco de dados encontrado: {db_path}")
        else:
            print(f"Erro: O banco de dados não foi encontrado no caminho {db_path}")

        # Carregar a imagem (ajustado para o diretório imgs)
        img_path = resource_path("imgs/atualizada.png")

        try:
            app_image_teste = Image.open(img_path)
            app_image_cadapp = Image.open(
                r"\\192.168.16.4\central de vendas\cadapp\.venv\cadapp_v.01\main_teste1\logo_cadzap-removebg-preview.png"
            )
            # print("Imagem carregada com sucesso.")

        except FileNotFoundError as e:
            print(f"Erro: Imagem não encontrada no caminho {img_path}")

        # Conectar ao banco de dados
        con = sqlite3.connect(db_path)

        # Diretório atual (onde o script está sendo executado)
        base_dir_atual = os.path.dirname(os.path.abspath(__file__))
        # print(f"Diretório atual: {base_dir_atual}")
        global tree
        global imagem_string, imagem, l_imagem

        # Função fechar app

        def fechar_app():
            janela_loggin.destroy()
            janela.destroy
            sys.exit()

        # DELETAR ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ

        # validar cpf e telefone
        def validar_campos():
            # Obter os valores dos campos
            telefone = e_tel.get().strip()
            cpf_cnpj = e_cpf_cnpj.get().strip()
            vendedor = e_vendedor.get().strip()
            status = e_status.get().strip()
            cidade = e_cidade.get().strip()
            bateria = e_bateria.get().strip()
            pagamento = e_forma_pgt.get().strip()
            base_troca = e_troca.get().strip()
            prazo = e_prazo.get().strip()
            midia = e_midia.get().strip()
            retorno = e_retorno.get().strip()
            canal = e_canal.get().strip()  # Canal deve ser obrigatório

            # Validar campos obrigatórios
            if not canal:
                messagebox.showerror("Erro", "O campo Canal é obrigatório.")
                return False

            # Validar telefone
            if telefone and len(telefone) != 11:
                messagebox.showerror(
                    "Erro", "O campo Telefone deve conter exatamente 11 números."
                )
                return False

            # Validar CPF/CNPJ
            if cpf_cnpj and len(cpf_cnpj) not in (11, 14):
                messagebox.showerror(
                    "Erro", "O campo CPF/CNPJ deve conter 11 ou 14 números."
                )
                return False

            # Validar campos com listas
            if vendedor and vendedor not in lista_vendedores:
                messagebox.showerror("Erro", f"O campo Vendedor deve ser um dos valores da lista")
                return False

            if status and status not in lista_status:
                messagebox.showerror("Erro", f"O campo Status deve ser um dos valores da lista")
                return False

            if cidade and cidade not in lista_cidades:
                messagebox.showerror("Erro", f"O campo Cidade deve ser um dos valores da lista")
                return False

            if bateria and bateria not in lista_baterias:
                messagebox.showerror("Erro", f"O campo Bateria deve ser um dos valores da lista")
                return False

            if pagamento and pagamento not in lista_forma_pgt:
                messagebox.showerror("Erro", f"O campo Forma de Pagamento deve ser um dos valores da lista")
                return False

            if base_troca and base_troca not in lista_troca:
                messagebox.showerror("Erro", f"O campo Base de Troca deve ser um dos valores da lista")
                return False

            if prazo and prazo not in lista_prazo:
                messagebox.showerror("Erro", f"O campo Prazo deve ser um dos valores da lista")
                return False

            if midia and midia not in lista_midias:
                messagebox.showerror("Erro", f"O campo Mídia deve ser um dos valores da lista")
                return False

            if retorno and retorno not in lista_retorno:
                messagebox.showerror("Erro", f"O campo Retorno deve ser um dos valores da lista")
                return False

            if canal and canal not in lista_canais:
                messagebox.showerror("Erro", f"O campo Canal deve ser um dos valores da lista")
                return False

            return True

        # validar valores
        def validar_valor_form():
            valor_form = e_valor_form.get().strip()

            # Verificar se o campo está vazio, se estiver, não precisa validar
            if not valor_form:
                return True  # Campo vazio é permitido

            # Tentar converter o valor para float (número real) se o campo não estiver vazio
            try:
                float(valor_form)
            except ValueError:
                messagebox.showerror(
                    "Erro", "O campo Valor/Form deve conter um número válido."
                )
                return False

            # Se for um número válido, retorna True
            return True

        # deletar os dados selecionados no banco de dados
        def deletar_form(i):
            with con:
                cur = con.cursor()
                ## na proxima coloca query para inserir os dados, dentro do VALUES, coloca ? de acordo com quantas colunas tem
                query_teste_deletar_dados = (
                    "DELETE FROM lancamentos_whatsapp_teste1 WHERE id = ?"
                )
                cur.execute(query_teste_deletar_dados, i)

        def deletar():
            global tree

            try:
                # Seleciona o item na árvore
                treev_dados = tree.focus()
                treev_dicionario = tree.item(treev_dados)
                treev_lista = treev_dicionario["values"]

                if not treev_lista:
                    messagebox.showwarning("Erro", "Selecione um item para deletar.")
                    return

                valor_id = treev_lista[0]  # Captura o ID do item a ser deletado

                # Função de confirmação para deletar
                def confirmar_deletar():
                    deletar_form([valor_id])  # Deletar o item do banco de dados
                    messagebox.showinfo("Sucesso", "Item deletado com sucesso.")
                    popupwindow.destroy()  # Fecha a janela de confirmação
                    mostrar()  # Atualiza a interface

                # Função para cancelar o processo de deletar
                def cancelar_deletar():
                    popupwindow.destroy()  # Fecha a janela de confirmação sem deletar

                # Criar a janela de confirmação
                popupwindow = Toplevel(janela)
                popupwindow.title("Confirmar Exclusão")
                popupwindow.geometry("250x150")

                # Centralizar a janela de confirmação
                largura_janela = 250
                altura_janela = 150
                largura_tela = popupwindow.winfo_screenwidth()
                altura_tela = popupwindow.winfo_screenheight()
                pos_x = (largura_tela // 2) - (largura_janela // 2)
                pos_y = (altura_tela // 2) - (altura_janela // 2)
                popupwindow.geometry(
                    f"{largura_janela}x{altura_janela}+{pos_x}+{pos_y}"
                )

                # Texto e botões de confirmação
                label_confirm = Label(
                    popupwindow, text="Deseja deletar este item?", font=("Arial", 10)
                )
                label_confirm.pack(pady=10)

                botao_sim = Button(popupwindow, text="Sim", command=confirmar_deletar)
                botao_sim.pack(pady=5)

                botao_nao = Button(
                    popupwindow, text="Cancelar", command=cancelar_deletar
                )
                botao_nao.pack(pady=5)

            except IndexError:
                messagebox.showerror("Erro", "Selecione um item para deletar.")

        # função inserir sql
        # Data de hoje
        hoje = datetime.date.today()
        data_formatada_hoje_exibir = hoje.strftime("%d/%m/%Y")
        # HORÁRIO DE PREENCHIMENTO
        agora = datetime.datetime.now()
        hora_para_exibir = agora.strftime(
            "%H:%M:%S"
        )  # Formato para exibição no formulário

        # Converter para o formato YYYY-MM-DD para o banco de dados
        data_para_inserir = hoje.strftime(
            "%Y-%m-%d"
        )  # Formato adequado para SQLite (YYYY-MM-DD)

        # Função inserir SQL com mapeamento correto para a planilha
        # Função inserir ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        def inserir():
            con = sqlite3.connect(
                r"\\192.168.16.4\central de vendas\cadapp\atendimento_wpp_teste1.db"
            )

            global user_entry  # Certifique-se de que user_entry está acessível
            user_info = user_entry.get()
            print("User info:", user_info)
            # Validar os campos de CPF/CNPJ, Telefone e Valor/Form
            if not validar_campos() or not validar_valor_form():
                return  # Se a validação falhar, interromper a inserção

            # Capturar os dados do formulário
            dados = {
                "VENDEDOR": e_vendedor.get(),
                "COD": e_cod.get(),
                "NOME": e_nome.get(),
                "TEL": e_tel.get(),
                "CPF/CNPJ": e_cpf_cnpj.get(),
                "STATUS": e_status.get(),
                "ENDEREÇO": e_endereco.get(),
                "CIDADE": e_cidade.get(),
                "LOCALIZAÇÃO": e_localizacao.get(),
                "REF": e_ref.get(),
                "CARRO/ANO": e_carro_ano.get(),
                "BATERIA": e_bateria.get(),
                "VALOR/FORM": e_valor_form.get(),
                "BASE DE TROCA": e_troca.get(),
                "PRAZO": e_prazo.get(),
                "MIDIA": e_midia.get(),
                "RETORNO": e_retorno.get(),
                "TABELA": e_tabela.get(),
                "CANAL": e_canal.get(),
                "QUEM IRÁ RECEBER": e_quem_ira_receber.get(),
                "OBS": e_obs.get(),
                "FORMA_PGT": e_forma_pgt.get(),
                "user_info": user_entry.get(),
            }

            # Colocar os dados em uma lista na ordem correta para a tabela no banco de dados
            lista_inserir = [
                dados["CANAL"],  # Canal
                "",  # Venda (deixe em branco por enquanto)
                dados["NOME"],  # Cliente (nome)
                dados["TEL"],  # Telefone
                dados["BATERIA"],  # Produto
                dados["VALOR/FORM"],  # Preço
                "",  # Loja (deixe em branco por enquanto)
                dados["VENDEDOR"],  # Atendente
                dados["STATUS"],  # Status
                dados["CPF/CNPJ"],  # CPF/CNPJ
                dados["MIDIA"],  # Mídia Social
                dados["RETORNO"],  # Retorno
                dados["CIDADE"],  # Cidade
                dados["CARRO/ANO"],  # Carro
                dados["OBS"],  # Observações
                "",  # Pós-venda (deixe em branco por enquanto)
                "",  # Motivo (deixe em branco por enquanto)
                "",  # Data de Nascimento (deixe em branco por enquanto)
                "",  # Família Produto (deixe em branco por enquanto)
                "",  # Mensagem (deixe em branco por enquanto)
                "",  # Enviar (definido como 'Não' por padrão)
                dados["COD"],  # COD
                dados["LOCALIZAÇÃO"],  # Localização
                dados["REF"],  # Referência
                dados["ENDEREÇO"],  # Endereço
                dados["BASE DE TROCA"],  # Base de Troca
                dados["PRAZO"],  # Prazo
                dados["TABELA"],  # Tabela
                dados["QUEM IRÁ RECEBER"],  # Quem Irá Receber
                "",  # DATA ATUALIZAÇÃO (deixe em branco por enquanto)
                "",  # HORA ATUALIZAÇÃO (deixe em branco por enquanto)
                dados["FORMA_PGT"],  # FORMA PGT (definido como 'Não' por padrão)
                dados["user_info"],
            ]

            try:
                # Inserir os dados na tabela de forma ordenada
                with con:
                    cur = con.cursor()
                    query = """
                        INSERT INTO lancamentos_whatsapp_teste1 (
                            data, canal, venda, cliente, telefone, hora, produto, preco, loja, atendente, status, cpf_cnpj, midia_social, 
                            retorno, cidade, carro, observacoes, pos_venda, motivo, dt_nascimento, familia_produto, mensagem, enviar, cod, 
                            localizacao, ref, endereco, base_troca, prazo, tabela, quem_ira_receber, data_atualizacao, hora_atualizacao, forma_pgt, user_info
                        ) VALUES (CURRENT_DATE, ?, ?, ?, ?, TIME('now', 'localtime'), ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?)
                    """
                    cur.execute(query, lista_inserir)
                    con.commit()

                # Mensagem de sucesso
                messagebox.showinfo("Sucesso", "Os dados foram inseridos com sucesso!")

            except sqlite3.Error as e:
                # Mensagem de erro caso ocorra algum problema na inserção
                messagebox.showerror(
                    "Erro", f"Ocorreu um erro ao inserir os dados: {e}"
                )

            finally:
                con.close()

            # Limpar os campos do formulário após a inserção
            campos_a_limpar = [
                e_nome,
                e_cod,
                e_tel,
                e_cpf_cnpj,
                e_status,
                e_endereco,
                e_cidade,
                e_localizacao,
                e_ref,
                e_carro_ano,
                e_bateria,
                e_valor_form,
                e_troca,
                e_prazo,
                e_midia,
                e_retorno,
                e_tabela,
                e_quem_ira_receber,
                e_obs,
                e_forma_pgt,
            ]

            for campo in campos_a_limpar:
                campo.delete(0, "end")

            # Atualizar a interface, se necessário
            mostrar()

        # ver os dados TODOS no banco de dados ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        def ver_dados_todos_form():
            con = sqlite3.connect(
                r"\\192.168.16.4\central de vendas\cadapp\atendimento_wpp_teste1.db"
            )
            cur = con.cursor()

            # Recuperar todos os dados de acordo com as colunas fornecidas no PRAGMA
            cur.execute(
                """
                SELECT id, data, hora, atendente, cod, cliente, telefone, cpf_cnpj, status, endereco, cidade, localizacao, ref, carro, produto, preco, base_troca, prazo, midia_social, retorno, tabela, canal, quem_ira_receber, observacoes 
                FROM lancamentos_whatsapp_teste1
                ORDER BY 
                SUBSTR(data, 7, 4) || '-' ||  -- Ano
                SUBSTR(data, 4, 2) || '-' ||  -- Mês
                SUBSTR(data, 1, 2) DESC,      -- Dia
                hora DESC                     -- Ordenar por hora também
            """
            )
            rows = cur.fetchall()
            con.close()
            return rows

        # ver os dados INDIVIDUALMENTE no banco de dados ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        def ver_dados_ind_form(id):

            ver_dados_individual = []
            with con:
                cur = con.cursor()
                ## na proxima coloca query para inserir os dados, dentro do VALUES, coloca ? de acordo com quantas colunas tem
                query_teste_ver_dados_individual = (
                    "SELECT * FROM lancamentos_whatsapp_teste1 WHERE id =?"
                )
                cur.execute(query_teste_ver_dados_individual, id)

                rows = cur.fetchall()
                for row in rows:
                    ver_dados_individual.append(row)

            return ver_dados_individual

        # funçãoo atualizar
        def atualizar():
            global imagem_string, imagem, l_imagem, b_avancar
            # Primeiro validar os campos

            try:
                # Seleciona o item na árvore e pega o ID
                treev_dados = tree.focus()  # Foco no item selecionado
                treev_dicionario = tree.item(treev_dados)  # Obter o dicionário do item
                treev_lista = treev_dicionario["values"]  # Lista de valores do item

                if not treev_lista:  # Verifica se um item foi realmente selecionado
                    raise IndexError

                # Captura o ID do item
                id = int(
                    treev_lista[0]
                )  # O ID do item a ser atualizado (chave primária)

                # Conectar ao banco de dados e obter os dados completos para o ID selecionado
                con = sqlite3.connect(db_path)
                cur = con.cursor()

                # Executar uma consulta para obter todos os dados do registro selecionado
                cur.execute(
                    "SELECT * FROM lancamentos_whatsapp_teste1 WHERE id = ?", (id,)
                )
                dados = cur.fetchone()

                # Verifica se dados foram encontrados
                if dados:
                    # Preencher os campos de entrada com os valores existentes
                    e_vendedor.delete(0, "end")
                    e_vendedor.insert(0, dados[10])  # Índice 10: 'atendente'
                    e_nome.delete(0, "end")
                    e_nome.insert(0, dados[4])  # Índice 4: 'cliente'
                    e_cod.delete(0, "end")
                    e_cod.insert(0, dados[24])  # Índice 24: 'cod'
                    e_tel.delete(0, "end")
                    e_tel.insert(0, dados[5])  # Índice 5: 'telefone'
                    e_cpf_cnpj.delete(0, "end")
                    e_cpf_cnpj.insert(0, dados[12])  # Índice 12: 'cpf_cnpj'
                    e_status.delete(0, "end")
                    e_status.insert(0, dados[11])  # Índice 11: 'status'
                    e_endereco.delete(0, "end")
                    e_endereco.insert(0, dados[27])  # Índice 27: 'endereco'
                    e_cidade.delete(0, "end")
                    e_cidade.insert(0, dados[15])  # Índice 15: 'cidade'
                    e_ref.delete(0, "end")
                    e_ref.insert(0, dados[26])  # Índice 26: 'ref'
                    e_carro_ano.delete(0, "end")
                    e_carro_ano.insert(0, dados[16])  # Índice 16: 'carro'
                    e_bateria.delete(0, "end")
                    e_bateria.insert(0, dados[7])  # Índice 7: 'produto'
                    e_valor_form.delete(0, "end")
                    e_valor_form.insert(0, dados[8])  # Índice 8: 'preco'
                    e_prazo.delete(0, "end")
                    e_prazo.insert(0, dados[29])  # Índice 29: 'prazo'
                    e_midia.delete(0, "end")
                    e_midia.insert(0, dados[13])  # Índice 13: 'midia_social'
                    e_retorno.delete(0, "end")
                    e_retorno.insert(0, dados[14])  # Índice 14: 'retorno'
                    e_tabela.delete(0, "end")
                    e_tabela.insert(0, dados[30])  # Índice 30: 'tabela'
                    e_quem_ira_receber.delete(0, "end")
                    e_quem_ira_receber.insert(
                        0, dados[31]
                    )  # Índice 31: 'quem_ira_receber'
                    e_obs.delete(0, "end")
                    e_obs.insert(0, dados[17])  # Índice 17: 'observacoes'
                    e_localizacao.delete(0, "end")
                    e_localizacao.insert(0, dados[25])  # Índice 25: 'localizacao'
                    e_canal.delete(0, "end")
                    e_canal.insert(0, dados[2])  # Índice 2: 'canal'
                    e_troca.delete(0, "end")
                    e_troca.insert(0, dados[28])  # Índice 28: 'base_troca'
                    e_forma_pgt.delete(0, "end")
                    e_forma_pgt.insert(0, dados[34])  # Índice 34: 'forma_pgt'

                con.close()

                # Função que realiza a atualização após a confirmação
                def update():
                    global imagem_string

                    # Capturar os novos valores dos campos
                    vendedor = e_vendedor.get()
                    nome = e_nome.get()
                    cod = e_cod.get()
                    tel = e_tel.get()
                    cpf_cnpj = e_cpf_cnpj.get()
                    status = e_status.get()
                    endereco = e_endereco.get()
                    cidade = e_cidade.get()
                    ref = e_ref.get()
                    carro_ano = e_carro_ano.get()
                    bateria = e_bateria.get()
                    valor_form = e_valor_form.get()
                    prazo = e_prazo.get()
                    midia = e_midia.get()
                    retorno = e_retorno.get()
                    tabela = e_tabela.get()
                    quem_ira_receber = e_quem_ira_receber.get()
                    obs = e_obs.get()
                    canal = e_canal.get()
                    localizacao = e_localizacao.get()
                    troca = e_troca.get()
                    forma_pgt = e_forma_pgt.get()
                    user_info = user_entry.get()

                    # Capturar a data e hora atuais para o registro de atualização
                    hoje = datetime.date.today()
                    agora = datetime.datetime.now()
                    data_para_inserir_update = hoje.strftime("%Y-%m-%d")
                    hora_para_inserir_update = agora.strftime("%H:%M:%S")

                    # Conectar ao banco de dados e atualizar o registro
                    con = sqlite3.connect(
                        r"\\192.168.16.4\central de vendas\cadapp\atendimento_wpp_teste1.db"
                    )
                    cur = con.cursor()
                    if not validar_campos():
                        return
                    if not validar_valor_form():
                        return
                    # Query para atualizar o registro com o ID específico
                    query_update = """
                        UPDATE lancamentos_whatsapp_teste1
                        SET atendente = ?, cliente = ?, cod = ?, telefone = ?, cpf_cnpj = ?, status = ?, 
                            endereco = ?, cidade = ?, localizacao = ?, ref = ?, carro = ?, produto = ?, 
                            preco = ?, base_troca = ?, prazo = ?, midia_social = ?, retorno = ?, tabela = ?, 
                            quem_ira_receber = ?, observacoes = ?, canal = ?, 
                            data_atualizacao = ?, hora_atualizacao = ?, forma_pgt = ? , user_info = ?
                        WHERE id = ?
                    """

                    cur.execute(
                        query_update,
                        (
                            vendedor,
                            nome,
                            cod,
                            tel,
                            cpf_cnpj,
                            status,
                            endereco,
                            cidade,
                            localizacao,
                            ref,
                            carro_ano,
                            bateria,
                            valor_form,
                            troca,
                            prazo,
                            midia,
                            retorno,
                            tabela,
                            quem_ira_receber,
                            obs,
                            canal,
                            data_para_inserir_update,
                            hora_para_inserir_update,
                            forma_pgt,
                            user_info,
                            id,
                        ),
                    )

                    con.commit()
                    con.close()

                    # Exibir mensagem de sucesso
                    messagebox.showinfo(
                        "Sucesso", "Os dados foram atualizados com sucesso"
                    )
                    b_avancar.destroy()
                    # Limpa os campos após a atualização
                    e_vendedor.delete(0, "end")
                    e_nome.delete(0, "end")
                    e_cod.delete(0, "end")
                    e_tel.delete(0, "end")
                    e_cpf_cnpj.delete(0, "end")
                    e_status.delete(0, "end")
                    e_endereco.delete(0, "end")
                    e_cidade.delete(0, "end")
                    e_ref.delete(0, "end")
                    e_carro_ano.delete(0, "end")
                    e_bateria.delete(0, "end")
                    e_valor_form.delete(0, "end")
                    e_prazo.delete(0, "end")
                    e_midia.delete(0, "end")
                    e_retorno.delete(0, "end")
                    e_tabela.delete(0, "end")
                    e_quem_ira_receber.delete(0, "end")
                    e_obs.delete(0, "end")
                    e_troca.delete(0, "end")
                    e_localizacao.delete(0, "end")
                    e_canal.delete(0, "end")
                    e_forma_pgt.delete(0, "end")

                    # Fecha o popup de confirmação
                    popupwindow.destroy()
                    b_avancar.destroy()
                    # Atualiza a interface
                    mostrar()

                # Função para abrir o popup de confirmação de atualização
                def popupatualizar():
                    global popupwindow
                    popupwindow = Toplevel(janela)
                    popupwindow.title("Atenção")

                    # Definir o tamanho da janela popup
                    largura_janela = 250
                    altura_janela = 150

                    # Obter o tamanho da tela
                    largura_tela = popupwindow.winfo_screenwidth()
                    altura_tela = popupwindow.winfo_screenheight()

                    # Calcular a posição x e y para centralizar a janela
                    pos_x = (largura_tela // 2) - (largura_janela // 2)
                    pos_y = (altura_tela // 2) - (altura_janela // 2)

                    # Definir a geometria da janela popup (tamanho e posição)
                    popupwindow.geometry(
                        f"{largura_janela}x{altura_janela}+{pos_x}+{pos_y}"
                    )

                    #comando para cancelar
                    def cancelar_update():
                 
                        # Limpa os campos após a atualização
                        e_vendedor.delete(0, "end")
                        e_nome.delete(0, "end")
                        e_cod.delete(0, "end")
                        e_tel.delete(0, "end")
                        e_cpf_cnpj.delete(0, "end")
                        e_status.delete(0, "end")
                        e_endereco.delete(0, "end")
                        e_cidade.delete(0, "end")
                        e_ref.delete(0, "end")
                        e_carro_ano.delete(0, "end")
                        e_bateria.delete(0, "end")
                        e_valor_form.delete(0, "end")
                        e_prazo.delete(0, "end")
                        e_midia.delete(0, "end")
                        e_retorno.delete(0, "end")
                        e_tabela.delete(0, "end")
                        e_quem_ira_receber.delete(0, "end")
                        e_obs.delete(0, "end")
                        e_troca.delete(0, "end")
                        e_localizacao.delete(0, "end")
                        e_canal.delete(0, "end")
                        e_forma_pgt.delete(0, "end")

                        popupwindow.destroy()
                        b_avancar.destroy()  
                        
                    Label(popupwindow, text="Confirmar atualização?", font=("verdana", 10)).pack(pady=10)
                    Button(popupwindow, text="Sim", command=update).pack(pady=5)
                    Button(popupwindow, text="Voltar", command=popupwindow.destroy).pack(pady=5)
                    Button(popupwindow, text="Cancelar", command=cancelar_update).pack(pady=5)


                # Criar o botão "Avançar" para permitir revisão dos dados antes de confirmar
                if "b_avancar" in globals():
                    b_avancar.destroy()  # Destruir o botão antigo antes de criar

                b_avancar = Button(
                    frame_esquerda,
                    command=popupatualizar,
                    width=11,
                    text="AVANÇAR".upper(),
                    overrelief=RIDGE,
                    font=("Ivy 12 bold"),
                    bg=cod_cor20,
                    fg=cod_cor16,
                )
                b_avancar.place(x=450, y=450, width=150)

            except IndexError:
                messagebox.showerror(
                    "Erro", "Selecione um dos dados na tabela para atualizar"
                )

        # limpar campos
        def limpar_campos():
            # Limpa os campos após a atualização
            e_vendedor.delete(0, "end")
            e_nome.delete(0, "end")
            e_cod.delete(0, "end")
            e_tel.delete(0, "end")
            e_cpf_cnpj.delete(0, "end")
            e_status.delete(0, "end")
            e_endereco.delete(0, "end")
            e_cidade.delete(0, "end")
            e_ref.delete(0, "end")
            e_carro_ano.delete(0, "end")
            e_bateria.delete(0, "end")
            e_valor_form.delete(0, "end")
            e_prazo.delete(0, "end")
            e_midia.delete(0, "end")
            e_retorno.delete(0, "end")
            e_tabela.delete(0, "end")
            e_quem_ira_receber.delete(0, "end")
            e_obs.delete(0, "end")
            e_troca.delete(0, "end")
            e_localizacao.delete(0, "end")
            e_canal.delete(0, "end")
            e_forma_pgt.delete(0, "end")


        def refresh():
            global mostrar
            mostrar()

        # Função para Selecionar o Intervalo de Datas
        import tkinter as tk
        from tkinter import messagebox, ttk

        import pandas as pd

        # Função para formatação automática da entrada de datas
        def format_date_entry(entry, next_entry=None):
            content = entry.get().replace("/", "")  # Remove barras durante a digitação
            if len(content) > 8:
                content = content[:8]  # Limita a 8 caracteres
            if len(content) > 4:
                entry.delete(0, tk.END)
                entry.insert(0, f"{content[:2]}/{content[2:4]}/{content[4:8]}")
            elif len(content) > 2:
                entry.delete(0, tk.END)
                entry.insert(0, f"{content[:2]}/{content[2:4]}")
            else:
                entry.delete(0, tk.END)
                entry.insert(0, content)

            # Verificar se a entrada tem 10 caracteres e mover para o próximo campo se necessário
            if len(entry.get()) == 10 and next_entry:  # Formato dd/mm/yyyy completo
                next_entry.focus_set()  # Move o foco para o próximo campo

        # Função para obter o intervalo de datas e o atendente (selecione o atendente de forma opcional)
        def get_date_range():
            def grab_dates():
                """Pega as datas e o atendente, e fecha a janela se estiverem válidos."""
                start_date = start_entry.get()
                end_date = end_entry.get()
                selected_atendente = combo_atendente.get()

                # Verificar o formato correto (dd/mm/yyyy)
                if len(start_date) != 10 or len(end_date) != 10:
                    messagebox.showwarning(
                        "Erro", "As datas devem estar no formato dd/mm/yyyy."
                    )
                    return None, None, None

                # Converter as datas para o formato SQL (yyyy-mm-dd)
                try:
                    start_date = pd.to_datetime(start_date, dayfirst=True).strftime(
                        "%Y-%m-%d"
                    )
                    end_date = pd.to_datetime(end_date, dayfirst=True).strftime(
                        "%Y-%m-%d"
                    )
                except ValueError:
                    messagebox.showerror(
                        "Erro", "Datas inválidas. Por favor, insira datas válidas."
                    )
                    return None, None, None

                root.destroy()  # Fechar a janela
                return start_date, end_date, selected_atendente

            def close_window():
                """Fecha a janela e retorna None para cancelar a operação."""
                root.destroy()  # Fecha a janela completamente
                return None, None, None

            def set_today():
                """Define a data de hoje para ambos os campos."""
                today = pd.to_datetime("today").strftime("%d/%m/%Y")
                start_entry.delete(0, tk.END)
                start_entry.insert(0, today)
                end_entry.delete(0, tk.END)
                end_entry.insert(0, today)

            def set_current_month():
                """Define o mês atual (início e fim do mês)."""
                first_day = pd.to_datetime("today").replace(day=1).strftime("%d/%m/%Y")
                last_day = (
                    pd.to_datetime("today").replace(day=1)
                    + pd.DateOffset(months=1)
                    - pd.DateOffset(days=1)
                )
                last_day = last_day.strftime("%d/%m/%Y")
                start_entry.delete(0, tk.END)
                start_entry.insert(0, first_day)
                end_entry.delete(0, tk.END)
                end_entry.insert(0, last_day)

            def set_current_year():
                """Define o ano atual (início e fim do ano)."""
                first_day_year = (
                    pd.to_datetime("today").replace(month=1, day=1).strftime("%d/%m/%Y")
                )
                last_day_year = (
                    pd.to_datetime("today")
                    .replace(month=12, day=31)
                    .strftime("%d/%m/%Y")
                )
                start_entry.delete(0, tk.END)
                start_entry.insert(0, first_day_year)
                end_entry.delete(0, tk.END)
                end_entry.insert(0, last_day_year)

            def set_all_until_today():
                """Define todas as datas desde o início até hoje."""
                first_date_db = "01/01/1900"
                today = pd.to_datetime("today").strftime("%d/%m/%Y")
                start_entry.delete(0, tk.END)
                start_entry.insert(0, first_date_db)
                end_entry.delete(0, tk.END)
                end_entry.insert(0, today)

            # Criar janela para o intervalo de datas e atendente
            root = tk.Toplevel()
            root.title("Selecione o intervalo de datas e o atendente (opcional)")

            # Centralizando a janela
            largura_janela = 450
            altura_janela = 300
            largura_tela = root.winfo_screenwidth()
            altura_tela = root.winfo_screenheight()
            pos_x = (largura_tela // 2) - (largura_janela // 2)
            pos_y = (altura_tela // 2) - (altura_janela // 2)
            root.geometry(f"{largura_janela}x{altura_janela}+{pos_x}+{pos_y}")

            # Layout para data inicial
            tk.Label(root, text="Data Inicial (dd/mm/yyyy):").grid(
                row=0, column=0, padx=10, pady=5
            )
            start_entry = tk.Entry(root, width=12)
            start_entry.grid(row=0, column=1)

            # Layout para data final
            tk.Label(root, text="Data Final (dd/mm/yyyy):").grid(
                row=1, column=0, padx=10, pady=5
            )
            end_entry = tk.Entry(root, width=12)
            end_entry.grid(row=1, column=1)

            # ComboBox para escolher o atendente
            tk.Label(root, text="Atendente (opcional):").grid(
                row=2, column=0, padx=10, pady=5
            )

            # Lista de atendentes (baseada em lista_vendedores existente)
            lista_vendedores = [
                "LUCAS AREDA - 98",
                "RAFAEL PATRICIO - 112",
                "CORPORATIVO - 100",
                "CLISTENES BRITO",
                "SANTIAGO SEGRE - 111",
                "DIOGO SANTANA - 113",
            ]

            combo_atendente = ttk.Combobox(root, values=lista_vendedores, width=25)
            combo_atendente.grid(row=2, column=1)
            combo_atendente.set("")  # Deixar o campo vazio inicialmente (opcional)

            # Botões de intervalo rápido
            tk.Button(root, text="Hoje", command=set_today).grid(
                row=0, column=2, padx=5, pady=5
            )
            tk.Button(root, text="Mês Atual", command=set_current_month).grid(
                row=1, column=2, padx=5, pady=5
            )
            tk.Button(root, text="Ano Atual", command=set_current_year).grid(
                row=2, column=2, padx=5, pady=5
            )
            tk.Button(root, text="Tudo até Hoje", command=set_all_until_today).grid(
                row=3, column=2, padx=5, pady=5
            )

            # Botão para confirmar
            confirm_button = tk.Button(
                root, text="Confirmar", command=lambda: root.quit()
            )
            confirm_button.grid(row=4, column=0, pady=10)

            # Botão para voltar
            back_button = tk.Button(root, text="Voltar", command=close_window)
            back_button.grid(row=4, column=1, pady=10)

            root.mainloop()

            return grab_dates()  # Retornar as datas e o atendente selecionados ou None

        # Função para buscar dados com base no intervalo de datas e atendente
        def fetch_data_by_date_range(
            db_path, table_name, start_date, end_date, atendente=None
        ):
            db_path = (
                r"\\192.168.16.4\central de vendas\cadapp\atendimento_wpp_teste1.db"
            )
            table_name = "lancamentos_whatsapp_teste1"

            try:
                # Conectar ao banco de dados
                conn = sqlite3.connect(db_path)

                # Query base com filtro de data
                query = f"""
                SELECT * FROM {table_name}
                WHERE DATE(data) BETWEEN '{start_date}' AND '{end_date}'
                """

                # Se um atendente for selecionado, adicionar filtro de atendente
                if atendente:
                    query += f" AND atendente = '{atendente}'"

                # Executar a query e carregar os dados em um DataFrame
                df = pd.read_sql_query(query, conn)

                if df.empty:
                    messagebox.showinfo(
                        "Sem dados",
                        "Nenhum dado encontrado para o intervalo selecionado.",
                    )
                    return None

                return df  # Retornar o DataFrame com os dados filtrados

            except Exception as e:
                messagebox.showerror(
                    "Erro", f"Erro ao tentar buscar os dados.\nDetalhes: {e}"
                )
                return None

            finally:
                conn.close()  # Fechar a conexão com o banco de dados

        # Função para escolher o caminho onde salvar o arquivo Excel
        def get_save_path():
            # Abrir uma caixa de diálogo para salvar o arquivo
            file_path = asksaveasfilename(
                defaultextension=".xlsx",  # Definir extensão padrão
                filetypes=[
                    ("Excel files", "*.xlsx"),
                    ("All files", "*.*"),
                ],  # Tipos de arquivo permitidos
                title="Salvar como",  # Título da caixa de diálogo
            )

            # Verificar se o caminho do arquivo foi selecionado ou o usuário cancelou
            if not file_path:
                messagebox.showwarning(
                    "Aviso", "Nenhum arquivo foi selecionado para salvar."
                )
                return None

            return file_path  # Retornar o caminho do arquivo selecionado

        # Função para salvar o DataFrame filtrado como arquivo Excel
        def save_data_to_excel(df):
            # Caminho para salvar o arquivo
            save_path = asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Salvar como",
            )

            if not save_path:
                messagebox.showwarning(
                    "Aviso", "Nenhum arquivo foi selecionado para salvar."
                )
                return

            # Criar um novo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dados Exportados"

            # Colunas organizadas na ordem correta
            colunas_ordenadas = [
                "data",
                "canal",
                "venda",
                "cliente",
                "telefone",
                "hora",
                "produto",
                "preco",
                "loja",
                "atendente",
                "status",
                "cpf_cnpj",
                "midia_social",
                "retorno",
                "cidade",
                "carro",
                "observacoes",
                "pos_venda",
                "motivo",
                "dt_nascimento",
                "familia_produto",
                "mensagem",
                "enviar",
            ]

            # Filtrar o DataFrame para manter apenas as colunas na ordem desejada
            df = df[colunas_ordenadas]

            # Adicionar o DataFrame ao Excel (mantendo a primeira linha como cabeçalho)
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), 1
            ):
                ws.append(row)

            # Aplicar os formatos diretamente como strings
            # 1. Data (formato dd/mm/yyyy)
            for cell in ws["A"][1:]:  # Coluna 'A' é 'data'
                cell.number_format = "DD/MM/YYYY"

            # 2. Hora (formato hh:mm:ss)
            for cell in ws["F"][1:]:  # Coluna 'F' é 'hora'
                cell.number_format = "HH:MM:SS"

            # 3. Telefone (sem formato especial, como texto)
            for cell in ws["E"][1:]:  # Coluna 'E' é 'telefone'
                cell.number_format = "@"  # Texto no Excel

            # 4. Preço (formato de moeda R$)
            for cell in ws["H"][1:]:  # Coluna 'H' é 'preco'
                cell.number_format = "R$ #,##0.00"  # Formato de moeda com R$

            # Salvar o workbook formatado
            try:
                wb.save(save_path)
                messagebox.showinfo(
                    "Sucesso",
                    "A planilha foi baixada com sucesso. \nLembre-se de colar como especial (ctrl + alt + v, apenas valores)",
                )
            except Exception as e:
                messagebox.showerror(
                    "Erro", f"Erro ao salvar o arquivo.\nDetalhes: {e}"
                )

        # Função Principal para o Download
        def download_table_with_date_filter(db_path, table_name):
            db_path = (
                r"\\192.168.16.4\central de vendas\cadapp\atendimento_wpp_teste1.db"
            )
            table_name = "lancamentos_whatsapp_teste1"

            # Obter o intervalo de datas e o atendente (opcional)
            start_date, end_date, atendente = get_date_range()

            # Verificar se as datas são válidas
            if not start_date or not end_date:
                return  # Se as datas não forem válidas, sair da função

            # Buscar os dados filtrados pelo intervalo de datas e, se houver, o atendente
            df = fetch_data_by_date_range(
                db_path, table_name, start_date, end_date, atendente
            )

            if df is not None and not df.empty:
                # Definir a ordem correta das colunas conforme a tabela fornecida
                colunas_ordenadas = [
                    "data",
                    "canal",
                    "venda",
                    "cliente",
                    "telefone",
                    "hora",
                    "produto",
                    "preco",
                    "loja",
                    "atendente",
                    "status",
                    "cpf_cnpj",
                    "midia_social",
                    "retorno",
                    "cidade",
                    "carro",
                    "observacoes",
                    "pos_venda",
                    "motivo",
                    "dt_nascimento",
                    "familia_produto",
                    "mensagem",
                    "enviar",
                ]

                # Verificar se todas as colunas existem no DataFrame antes de ordenar
                colunas_disponiveis = [
                    col for col in colunas_ordenadas if col in df.columns
                ]
                df_ordenado = df[
                    colunas_disponiveis
                ]  # Reorganizar conforme as colunas disponíveis

                # Formatar a coluna de data para o formato dd/mm/aaaa
                if "data" in df_ordenado.columns:
                    df_ordenado["data"] = pd.to_datetime(
                        df_ordenado["data"], format="%Y-%m-%d"
                    ).dt.strftime("%d/%m/%Y")

                # Chamar a função para salvar o DataFrame em um arquivo Excel
                save_data_to_excel(df_ordenado)
            else:
                messagebox.showinfo(
                    "Sem dados",
                    "Nenhum dado foi encontrado para o intervalo selecionado.",
                )

        # Função para capturar os dados do formulário e formatar a mensagem
        def gerar_mensagem_formatada():
            try:
                # Capturar a linha selecionada na tabela
                treev_dados = tree.focus()  # Foco na linha selecionada
                treev_dicionario = tree.item(treev_dados)  # Obter o dicionário do item
                treev_lista = treev_dicionario["values"]  # Lista de valores do item

                if not treev_lista:  # Verifica se um item foi realmente selecionado
                    raise IndexError

                # Captura o ID da linha selecionada (assumindo que a coluna 0 contém o ID)
                id_selecionado = treev_lista[0]

                # Conectar ao banco de dados e buscar os dados da linha selecionada
                con = sqlite3.connect(db_path)
                with con:
                    cur = con.cursor()
                    # Consultar os dados da linha selecionada com base no ID
                    cur.execute(
                        "SELECT * FROM lancamentos_whatsapp_teste1 WHERE id = ?",
                        (id_selecionado,),
                    )
                    linha = cur.fetchone()  # Obter a linha correspondente ao ID

                # Se a linha for encontrada, extrair os dados
                if linha:
                    # Atribuindo os valores da linha selecionada para as variáveis correspondentes
                    id = linha[0]  # id tabela (coluna 0)
                    data = linha[1]  # Data da compra (coluna 1)
                    hora = linha[6]  # Hora da compra (coluna 6)
                    vendedor = linha[10]  # Atendente/Vendedor (coluna 10)
                    cod = linha[24]  # Venda/Código (coluna 24)
                    nome = linha[4]  # Nome do cliente (coluna 4)
                    telefone = linha[5]  # Telefone (coluna 5)
                    cpf_cnpj = linha[12]  # CPF/CNPJ (coluna 12)
                    endereco = linha[27]  # Endereço (coluna 27)
                    cidade = linha[15]  # Cidade (coluna 15)
                    localizacao = linha[25]  # Localização (coluna 25)
                    ref = linha[26]  # Produto/Referência (coluna 26)
                    carro_ano = linha[16]  # Carro/Ano (coluna 16)
                    bateria = linha[7]  # Produto/Bateria (coluna 7)
                    valor = linha[8]  # Valor (coluna 8)
                    prazo = linha[29]  # Prazo (coluna 29)
                    base_troca = linha[28]  # Base de troca (coluna 28)
                    midia = linha[13]  # Mídia social (coluna 13)
                    tabela = linha[30]  # Tabela (coluna 30)
                    quem_receber = linha[31]  # Quem vai receber (coluna 31)
                    obs = linha[17]  # Observações (coluna 17)
                    status = linha[11]  # Status (coluna 11)
                    retorno = linha[14]  # Retorno (coluna 14)
                    forma_pgt = linha[34]

                    # Formatar a mensagem sem tabs extras
                    mensagem_formatada = f"""
*VENDEDOR:* {vendedor}
*DATA/HORÁRIO:* {data} {hora}

*COD:* {cod}
*CLIENTE:* {nome}
*TELEFONE:* {telefone}
*CPF/CNPJ:* {cpf_cnpj}

*ENDEREÇO:* {endereco}
*CIDADE:* {cidade}
*LOCALIZAÇÃO:* {localizacao}
*REF:* {ref}
                    
*CARRO/ANO:* {carro_ano}
*BATERIA:* {bateria}
*VALOR:* {valor} 
*F. PAGAMENTO:* {forma_pgt}
*BASE DE TROCA:" {base_troca}
*PRAZO:* {prazo}

*TABELA:* {tabela}
*QUEM IRÁ RECEBER?* {quem_receber}

*OBS:* {obs}"""

                    # Copiar a mensagem para o clipboard
                    import pyperclip

                    pyperclip.copy(mensagem_formatada.strip())

                    # Exibir uma mensagem de confirmação
                    messagebox.showinfo(
                        "Sucesso",
                        "A mensagem foi gerada e copiada para o clipboard.\nAgora você pode colar no WhatsApp.",
                    )

                else:
                    messagebox.showerror(
                        "Erro",
                        "Não foi possível encontrar os dados da linha selecionada.",
                    )

            except IndexError:
                messagebox.showerror(
                    "Erro", "Selecione uma linha da tabela para gerar a mensagem."
                )

        # Função que converte o texto do Entry para maiúsculas
        def converter_maiusculas(event):
            # Verificação simples para debugar
            entrada = event.widget  # Obter o widget Entry onde o evento ocorreu
            posicao_cursor = entrada.index(tk.INSERT)  # Posição atual do cursor
            texto_atual = entrada.get()  # Texto atual no campo Entry
            texto_maiusculo = texto_atual.upper()  # Converte o texto para maiúsculas

            # Atualizar o texto para maiúsculas somente se houver diferença
            if texto_atual != texto_maiusculo:
                entrada.delete(0, tk.END)  # Limpar o texto atual
                entrada.insert(0, texto_maiusculo)  # Inserir o texto em maiúsculas
                entrada.icursor(posicao_cursor)  # Reposicionar o cursor

            # Ajustar a visualização para centralizar o cursor
            comprimento_texto = len(texto_maiusculo)
            if comprimento_texto > 0:
                # Definir intervalo de 3 caracteres antes e depois do cursor para visualização
                inicio_visivel = max(0, posicao_cursor - 3)
                fim_visivel = min(comprimento_texto, posicao_cursor + 3)

                # Calcular o deslocamento centralizado para manter o cursor no campo visível
                posicao_centralizada = inicio_visivel / comprimento_texto
                entrada.xview_moveto(posicao_centralizada)

        # Criando Janela vazia ###############################################################################################################################

        # Função para ajustar o tamanho da janela com base na tela do usuário
        def ajustar_janela():
            global altura_janela, largura_janela

            largura_tela = janela.winfo_screenwidth()  # Largura da tela
            altura_tela = janela.winfo_screenheight()  # Altura da tela

            largura_janela = int(largura_tela * 0.8)  # 80% da largura da tela
            altura_janela = int(altura_tela * 0.8)  # 80% da altura da tela

            pos_x = (largura_tela // 2) - (
                largura_janela // 2
            )  # Centraliza horizontalmente
            pos_y = (altura_tela // 2) - (
                altura_janela // 2
            )  # Centraliza verticalmente

            janela.geometry(f"{largura_janela}x{altura_janela}+{pos_x}+{pos_y}")
            janela.configure(bg="#EFEFEF")
            janela.resizable(True, True)  # Janela ajustável

        # Criando Janela
        janela = Tk()
        janela.title("CadZap - Lançamentos Whatsapp")
        ajustar_janela()
        # estilo da janela
        style = ttk.Style(janela)
        style.theme_use("clam")

        # Configurar a linha do cabeçalho (frame_cima) para se ajustar dinamicamente
        janela.grid_rowconfigure(0, weight=0)  # O cabeçalho só vai à esquerda
        janela.grid_rowconfigure(1, weight=1)  # O conteúdo principal tem peso maior
        janela.grid_columnconfigure(
            0, weight=1, uniform="metade"
        )  # Para a parte esquerda
        janela.grid_columnconfigure(
            1, weight=1, uniform="metade"
        )  # Para a parte direita

        # FRAME CABEÇALHO (Somente na metade esquerda)
        frame_cima = Frame(
            janela, bg=cod_cor15, relief=FLAT
        )  # Ajuste da altura para ser mais compacto
        frame_cima.grid(row=0, column=0, sticky="nsew")  # Cabeçalho na metade esquerda
        try:

            # Carregar a imagem (ajuste o caminho se necessário)
            app_image_cadapp = Image.open(
                resource_path("logo_cadzap-removebg-preview.png")
            )
            # print('logo carregado')
            app_image_cadapp = app_image_cadapp.resize(
                (40, 40)
            )  # Diminuir o tamanho da imagem
            # print('imagem redimensionada logo')
            app_image_cadapp = ImageTk.PhotoImage(app_image_cadapp)
            # print('imagem convertida logo')

        except Exception as e:
            print(f"Ocorreu um erro: {e}")

        try:
            # Adicionar logotipo no frame_cima
            label_logo = Label(frame_cima, bg=cod_cor15)
            # print('logo label')
            label_logo.pack(side=LEFT)
        except Exception as e:
            print(f"Ocorreu um erro: {e}")

        # Definir fonte moderna e mais compacta
        modern_font = Font(family="Segoe UI", size=18, weight="bold")

        # Adicionar Texto do Cabeçalho (centralizado e ajustado)
        label_framecima = Label(
            frame_cima,
            text="CADAPP - Cadastro de Vendas no WhatsApp",
            bg=cod_cor15,
            fg=cod_cor01,
            font=modern_font,
            height=1,
        )
        label_framecima.pack(side=LEFT)

        # Criar um frame para a versão (à direita do cabeçalho)
        frame_versao = Frame(frame_cima, bg=cod_cor15)
        frame_versao.pack(side=BOTTOM)

        # Adicionar rótulo no frame_versao
        label_versao = Label(
            frame_versao,
            text="Versão 1.0",
            bg=cod_cor15,
            fg="gray",
            font=("Segoe UI", 10),
        )
        label_versao.pack(side=BOTTOM)

        



        # FRAME ESQUERDA - Com barras de rolagem
        # Criar um Canvas para o frame esquerdo
        canvas_esquerda = Canvas(janela, bg=cod_cor01)

        # Configuração para ocupar 50% da largura da tela
        canvas_esquerda.grid(row=1, column=0, sticky="nsew")  # Canvas vai expandir
        janela.grid_columnconfigure(
            0, weight=1
        )  # A primeira coluna ocupa metade da janela
        janela.grid_columnconfigure(
            1, weight=1
        )  # A segunda coluna ocupa a outra metade da janela
        janela.grid_rowconfigure(
            1, weight=1
        )  # Permitir que o conteúdo se expanda na linha também

        # Barra de rolagem vertical
        scrollbar_esquerda_vertical = Scrollbar(
            janela, orient=VERTICAL, command=canvas_esquerda.yview
        )
        scrollbar_esquerda_vertical.grid(row=1, column=0, sticky="nse")

        # Barra de rolagem horizontal
        scrollbar_esquerda_horizontal = Scrollbar(
            janela, orient=HORIZONTAL, command=canvas_esquerda.xview
        )
        scrollbar_esquerda_horizontal.grid(row=2, column=0, sticky="ew")

        # Configurar o canvas para usar as barras de rolagem
        canvas_esquerda.configure(
            yscrollcommand=scrollbar_esquerda_vertical.set,
            xscrollcommand=scrollbar_esquerda_horizontal.set,
        )

        # Criar um frame dentro do canvas
        frame_esquerda = Frame(canvas_esquerda, bg=cod_cor01)
        canvas_esquerda.create_window((0, 0), window=frame_esquerda, anchor="nw")

        # Função para ajustar a rolagem conforme o conteúdo do frame muda
        def on_frame_configure(event):
            canvas_esquerda.configure(scrollregion=canvas_esquerda.bbox("all"))

        # Vincular o evento de configuração do frame ao canvas
        frame_esquerda.bind("<Configure>", on_frame_configure)

        # Função para habilitar a rolagem horizontal e vertical com o mouse
        def scroll_with_mouse(event):
            if event.state == 0:  # Sem modificadores
                canvas_esquerda.yview_scroll(int(-1 * (event.delta / 120)), "units")
            elif event.state == 1:  # Shift pressionado
                canvas_esquerda.xview_scroll(int(-1 * (event.delta / 120)), "units")

        # Bind para rolar com o mouse (scroll vertical e horizontal)
        canvas_esquerda.bind_all("<MouseWheel>", scroll_with_mouse)

        # FRAME DIREITA - Dividido em cima e baixo
        frame_direita = Frame(janela, bg="#A0A0A0", relief=FLAT)
        frame_direita.grid(
            row=0, column=1, rowspan=2, sticky="nsew"
        )  # Frame da direita (50% da largura, toda a altura)

        # Dividindo o frame_direita em frame_direita_cima e frame_direita_baixo
        frame_direita_cima = Frame(frame_direita, bg="#A0C0A0", relief=FLAT)
        frame_direita_cima.pack(fill=BOTH, expand=True)  # Ocupará a metade superior
        

        frame_direita_baixo = Frame(frame_direita, bg=cod_cor04, relief=FLAT)
        frame_direita_baixo.pack(fill=BOTH, expand=True)  # Ocupará a metade inferior

        # Configurar redimensionamento horizontal para os frames esquerda e direita
        janela.grid_columnconfigure(
            0, weight=1, uniform="metade"
        )  # Define que ambos frames ocupam metade
        janela.grid_columnconfigure(1, weight=1, uniform="metade")

        ###### CRIANDO LISTAS #####################################################################################
        # Criar a lista de vendedores
        lista_canais = ["AUTO BATERIAS", "GUGA TAGUA"]
        
        lista_vendedores = [
                "LUCAS AREDA - 98",
                "RAFAEL PATRICIO - 112",
                "CORPORATIVO - 100",
                "CLISTENES BRITO",
                "SANTIAGO SEGRE - 111",
                "DIOGO SANTANA - 113",
                "THALES FORTES - 97",
            ]

        # Lista Status
        lista_status = [
            "CANCELADA",
            "ENGANO",
            "GARANTIA",
            "INFORMAÇAO",
            "NAO",
            "SEM ESTOQUE",
            "SOCORRO",
            "TRANSFERENCIA",
            "VENDIDA",
            "SIM",
            "SEM PRODUTO",
            "PEÇAS",
            "NOTA FISCAL",
            "MOTO",
            "FREEDOM",
        ]

        # Lista de cidades
        lista_cidades = [
            "AGUAS CLARAS",
            "AGUAS LINDAS",
            "AREAL",
            "ARNIQUEIRAS",
            "ASA NORTE",
            "ASA SUL",
            "BRAZLANDIA",
            "CANDANGOLANDIA",
            "CEILANDIA",
            "CRUZEIRO",
            "ESTRUTURAL",
            "FORMOSA",
            "GAMA",
            "GRANJA DO TORTO",
            "GUARA",
            "JARDIM BOTANICO",
            "JARDIM INGA",
            "LAGO NORTE",
            "LAGO OESTE",
            "LAGO SUL",
            "LUZIANIA",
            "MANGUEIRAL",
            "NOROESTE",
            "NUCLEO BANDEIRANTE",
            "OCIDENTAL",
            "OCTOGONAL",
            "PARANOA",
            "PARK WAY",
            "PLANALTINA",
            "PLANALTINA DE GOIAS",
            "PLANO PILOTO",
            "RECANTO DAS EMAS",
            "RIACHO FUNDO",
            "S.CLUBES NORTE",
            "S.CLUBES SUL",
            "SAAN",
            "SAMAMBAIA",
            "SANTA MARIA",
            "SAO SEBASTIAO",
            "SIA",
            "SOBRADINHO",
            "SOF NORTE",
            "SOF SUL",
            "SUDOESTE",
            "TAGUATINGA",
            "VALPARAISO",
            "VARJAO",
            "VICENTE PIRES",
            "VILA PLANALTO",
        ]

        # Lista de Mídias
        lista_midias = ["WHATSAPP", "GOOGLE"]

        # Lista de trocas
        lista_troca = ["A BASE DE TROCA", "SEM TROCA"]

        # Lista de prazo
        lista_prazo = ["50 MINUTOS", "URGENTEE!"]

        lista_baterias = [
            "HFS150TD",
            "HFT150TD",
            "HFT180TD",
            "HFS180TD",
            "HG38JD",
            "HG40FD",
            "HG45BD",
            "HG45BE",
            "HG45JE",
            "HG50GD",
            "HG50JD",
            "HG70ND",
            "HG70NE",
            "HG70PD",
            "HG75LD",
            "HG75LE",
            "HG90LD",
            "HG90LE",
            "HG95MD",
            "HGD75LD",
            "HGR60DD",
            "HGR60DE",
            "HNP45BD",
            "HNP50GD",
            "HNP60DD",
            "HNP60DE",
            "HNP60HD",
            "RT100LE",
            "RT150TD",
            "RT200TD",
            "RT225TE",
            "RTP100LE",
            "SRT180TD",
            "SRT180TE",
            "AG60HD",
            "AG70PD",
            "AG80KD",
            "AG95MD",
            "HAG70PD",
            "HAG92MD",
            "AG105SD",
            "HF48BD",
            "HF52GD",
            "HF60DD",
            "HF60HD",
            "HF65HD",
            "HF70ND",
            "HF75PD",
            "HFB60HD",
            "HFB72PD",
            "HV48BD",
            "HV52GD",
            "HV60DD",
            "HV60HD",
            "AM100LE",
            "AM170TD",
            "AM38JD",
            "AM45JD",
            "AMR45JD",
            "AM75LD",
            "AM90LD",
            "AMF150TD",
            "AMR45BD",
            "AMR45BE",
            "AMR50GD",
            "AMR60DD",
            "AMR60DE",
            "AMR60HD",
            "AMR70ND",
            "AMR70NE",
            "ST100LE",
            "ST150TD",
            "ST45BD",
            "ST50GD",
            "ST60DD",
            "ST60DE",
            "ST60HD",
            "ST70ND",
            "ST70NE",
            "HT12ABS",
            "HTX14",
            "HTX16BS",
            "HTX16BS",
            "HTX7ABS",
            "HTX7A-BS",
            "HTX9BS",
            "HTX9-BS",
            "HTZ10SBS",
            "HTZ10S-BS",
            "HTZ14SBS",
            "HTZ12SBS",
            "HTZ5L",
            "HTX12BS",
            "HTZ6L",
            "HTZ7L",
            "ATX5L",
            "ATX6L",
            "ATX7L",
            "DF 12-18",
            "DF1000",
            "DF12-12",
            "DF12-18",
            "DF12-7",
            "DF12-7E",
            "DF12-9",
            "DF2000",
            "DF2500",
            "DF300",
            "DF3000",
            "DF4100",
            "DF500",
            "DF700",
            "HN100",
            "HN150",
            "HN45",
            "HN60",
            "HTX12-BS",
            "H38JD",
            "H40FD",
            "H45JE",
            "H48BD",
            "H52GD",
            "H50JD",
            "H60DD",
            "H60HD",
            "H65HD",
            "H70ND",
            "H75LD",
            "H75PD",
            "HD75LD",
            "HD90LD",
            "H90LD",
            "H95MD",
            "HEFB60HD",
            "HEFB72PD",
            "HEFB50GD",
            "HAGM60HD",
            "HAGM70PD",
            "HAGM80KD",
            "HAGM95MD",
            "HAGM105SD",
            "H100LE",
            "H150TD",
            "H180TD",
            "H180TE",
            "HS200TD",
            "HEFB225TE",
            "HS100LE",
            "HS150TD",
            "HS180TD",
            "H70NE",
            "H90LE",
            "RT150TD",
            "E45BD",
            "E50GD",
            "E60DD",
            "E60DE",
            "E60HD",
            "E70ND",
            "E70NE",
            "E75PD",
            "E75LD",
            "E90LD",
            "E95MD",
            "E100LE",
            "E150TD",
            "E180TD",
            "EFB50GD",
            "EFB60HD",
            "EFB72PD",
            "EGM60HD",
            "EGM70PD",
            "EGM80KD",
            "EGM92MD",
            "E42JD",
            "E52JD",
            "EGM60HD",
            "EGM70PD",
            "EGM80KD",
            "EGM92MD",
            "H60DE",
            "H75LE",
        ]

        # Lista Retorno
        lista_retorno = [
            "ACHOU O PREÇO ALTO",
            "ACIONOU O SEGURO",
            "QUERIA URGENTE",
            "CHAMADA NÃO COMPLETADA",
            "COMPROU AQUI EM OUTRA DATA",
            "COMPROU COM OUTRO ATENDENTE",
            "COMPROU MOURA MAIS EM CONTA",
            "COMPROU NA LOJA PRÓXIMO A ELE",
            "CONCORRENTE",
            "CONSEGUIU LIGAR O CARRO",
            "COTAÇÃO",
            "NÃO VAI COMPRAR",
            "VAI NA AUTORIZADA",
            "FORA DE APLICAÇÃO",
            "QUER MAIS EM CONTA",
            "JÁ ACHOU MAIS EM CONTA",
            "LOJISTA - MECAN",
            "MARCA ZETA",
            "NÃO ATENDEU",
            "NÃO ERA BATERIA - O CARRO PEGOU",
            "NÃO QUIS INFORMAR",
            "NEGOCIANDO PELO WHATSAPP",
            "NÚMERO INCORRETO",
            "NÚMERO INVÁLIDO",
            "OUTRO ESTADO",
            "OUTROS",
            "PARENTE JÁ TINHA COMPRADO",
            "PESQ. PARA TERCEIROS",
            "PESQUISA PREVENTIVA",
            "PROCURA DE 3ª LINHA",
            "QUERIA ACDELCO",
            "ERRO LOGÍSTICA",
            "QUERIA CRAL",
            "QUERIA MOURA",
            "QUERIA PIONEIRO",
            "RETORNOU E COMPROU",
            "SEM LOJA",
            "TELEFONE ERRADO",
            "VAI COMPRAR DEPOIS",
            "VAI NA LOJA",
            "WHATSAPP",
            "BATERIA EM GARANTIA",
            "FALTA DE BOY",
            "ATRASO",
            "CLIENTE DESISTIU",
            "NÃO ERA BATERIA",
            "ATRASO DE AGENDAMENTO",
            "CARRO PEGOU",
            "PROBLEMA COM CARTÃO",
            "CLIENTE NÃO ESTAVA NO LOCAL",
            "SEM PRODUTO NA LOJA",
            "NÃO TINHA SUCATA",
            "ESTACIONÁRIA",
            "CLIENTE DENTRO DE OUTRA LOJA DE BATERIA FAZENDO LEILÃO",
            "LOJA NÃO ENVIOU",
            "NÃO RESPONDEU",
            "BOY NÃO LEVOU MULTÍMETRO",
            "PESQUISANDO",
            "APLICAÇÃO INCORRETA",
            "PREÇO SEM TROCA - ACHOU CARO",
            "ENTREGA EM STO ANTONIO DESC COM URGÊNCIA",
            "NÃO RESPONDEU",
            "CHAMOU O SEGURO",
            "CLIENTE PEDIU EM 2 LOJAS",
            "PENDÊNCIA DE SUCATA",
            "HF48BD",
            "HF52GD",
            "HF60DD",
            "HF60HD",
            "HF65HD",
            "HF70ND",
            "HF75PD",
            "HFB60HD",
            "HG38JD",
            "HG40FD",
            "HG45BD",
            "HG45BE",
            "HG45JE",
            "HG50GD",
            "HG50JD",
            "HG60HD",
            "HG70ND",
            "HG70NE",
            "HG70PD",
            "HG75LD",
            "HG75LE",
            "HG90LD",
            "HG90LE",
            "HG95MD",
            "HGD75LD",
            "HGD90LD",
            "HGR60DD",
            "HGR60DE",
            "RT100LE",
            "RTV150TD",
            "RTV170TD",
            "SRT180TD",
            "SRT180TE",
            "RT200TD",
            "RT225TE",
            "RTP100LE",
            "RTP150TD",
            "RTP170TD",
            "HFT150TD",
            "HFT180TD",
            "AG60HD",
            "AG70PD",
            "AG80KD",
            "AG95MD",
            "AG105SD",
            "AM38JD",
            "AM45BD",
            "AM45BE",
            "AM45JD",
            "AM50DD",
            "AM50GD",
            "AM60DD",
            "AM60DE",
            "AM60HD",
            "AM70ND",
            "AM70NE",
            "AM75LD",
            "AM90LD",
            "AM100LE",
            "AM150TD",
            "AM170TD",
            "AMS180TE",
            "AMR60DE",
            "AMR70NE",
            "FRETE",
            "DF12-4",
            "DF12-7E",
            "DF12-7",
            "DF12-9",
            "DF12-12",
            "DF12-18",
            "REPARO",
            "PL50D",
            "PL50E",
            "PL90FE",
            "HFS150TD",
            "DF500",
            "DF700",
            "DF1000",
            "DF1500",
            "DF2000",
            "DF2500",
            "DF3000",
            "DF4100",
            "DF6-12",
            "DF6-4",
            "ATX5L",
            "ATX6L",
            "ATX7L",
            "HTZ5L",
            "HTZ6L",
            "HTZ7L",
            "ST50GD",
            "HTX9BS",
            "HTZ10SBS",
            "HTZ12SBS",
            "ST70NE",
            "HTX12BS",
            "HTZ14SBS",
            "HTX16BS",
            "ST45BE",
            "HT12ABS",
            "ST100LE",
            "HN45",
            "HN60",
            "HTX7ABS",
            "HN100",
            "HN150",
            "HTX14",
            "AMR45BE",
            "AMF150TD",
            "HNP60DD",
            "DF300",
            "AMR50GD",
            "AMR60DD",
            "AMR60HD",
            "HFB72PD",
            "HNP45BD",
            "HNP50GD",
            "HNP60HD",
            "HTX7ABS1",
            "ST45BD",
            "ST60DD",
            "ST60HD",
            "AMR45BD",
            "AMR70ND",
            "ST150TD",
            "ST70ND",
            "ST60DE",
            "HFS180TD",
            "HV48BD",
            "HV52GD",
            "HV60DD",
            "HV60HD",
            "DF12-7.2",
            "ESTACIONÁRIA",
            "CORPORATIVO",
            "FAVOR, COLOCAR O PRODUTO AQUI!!!!!",
            "QUAL PRODUTO ESTÁ EM FALTA???????",
            "TRANSFERENCIA",
            "LIGAÇÃO CAIU",
            "LIGAÇÃO MUDA",
            "GRAVAÇÃO",
            "SEM ESTOQUE - MAIS EM CONTA",
            "CLIENTE DE OUTRO VENDEDOR",
        ]

        lista_forma_pgt = [
            "CREDITO A VISTA",
            "2x",
            "3x",
            "4x",
            "5x",
            "6x",
            "DEBITO",
            "PIX",
            "BOLETO",
        ]

        # FRAME ESQUERDA XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX
        ## Trabalhando no Frame do meio ---------------------------------------
        # Título "Vendedor"zzzzzzzzzzzzzzzzzzzzzzzzzzzzzzz
        l_vendedor = Label(
            frame_esquerda,
            text="VENDEDOR:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_vendedor.grid(row=0, column=0, padx=10, pady=5, sticky=W)
        e_vendedor = ttk.Combobox(frame_esquerda, values=lista_vendedores, width=30)
        e_vendedor.grid(row=0, column=1, padx=10, pady=5, sticky=W)

        # Habilitar o preenchimento automático conforme você digita
        e_vendedor["state"] = "normal"
        e_vendedor.set("")  # Limpar o campo inicialmente
        # Definir a lista de opções inicialmente
        e_vendedor["values"] = lista_vendedores

        # Função para mostrar a lista suspensa sem travar o campo de entrada
        def show_dropdown(event):
            typed_text = e_vendedor.get()

            # Apenas exibir a lista de sugestões, sem modificar o que está digitado
            if len(typed_text) >= 3:
                # Filtrar vendedores que contenham a sequência digitada
                filtered_values = [
                    v for v in lista_vendedores if typed_text.lower() in v.lower()
                ]
                if filtered_values:
                    e_vendedor["values"] = filtered_values  # Exibir sugestões filtradas
                    e_vendedor.event_generate("<Down>")  # Mostrar a lista suspensa
                else:
                    e_vendedor["values"] = (
                        []
                    )  # Se não houver correspondências, esvaziar a lista
            else:
                e_vendedor["values"] = (
                    []
                )  # Limpar a lista suspensa se houver menos de 3 caracteres

        # Validação ao sair do campo, apenas aceita valores da lista
        def validate_choice(event=None):
            typed_text = e_vendedor.get()
            if typed_text not in lista_vendedores:
                pass

        # Função para autocompletar com o "Tab", sem interferir na digitação normal
        def auto_complete(event):
            typed_text = e_vendedor.get()
            if typed_text == "":
                e_vendedor["values"] = lista_vendedores
            else:
                # Filtra a lista de acordo com o que foi digitado
                suggestions = [
                    v for v in lista_vendedores if typed_text.lower() in v.lower()
                ]
                if suggestions:
                    e_vendedor.set(suggestions[0])  # Completa com a primeira sugestão
                    e_vendedor.icursor(
                        len(typed_text)
                    )  # Mantém o cursor após o texto digitado
                    pass

        # Bind para mostrar sugestões conforme o usuário digita
        e_vendedor.bind("<FocusOut>", validate_choice)
        # Bind para mostrar sugestões conforme o usuário digita
        e_vendedor.bind("<KeyRelease>", show_dropdown)
        e_vendedor.bind("<KeyRelease>", converter_maiusculas)
        # Evento para completar com "Tab"
        e_vendedor.bind("<Tab>", auto_complete, validate_choice)
        # Evento para completar com "Tab"
        e_vendedor.bind("<Enter>", auto_complete, validate_choice)

        # Definir a lista de opções inicialmente e limpar o campo
        e_vendedor["state"] = "normal"
        e_vendedor.set("")

        # Título "Cod" zzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzz
        l_cod = Label(
            frame_esquerda,
            text="COD:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_cod.grid(row=1, column=0, padx=10, pady=5, sticky=W)
        e_cod = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=10)
        e_cod.grid(row=1, column=1, padx=10, pady=5, sticky=W)
        # Vincular o evento de pressionamento de teclas para converter o texto em maiúsculas
        e_cod.bind("<KeyRelease>", converter_maiusculas)
        # Título "Nome" zzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzz
        l_nome = Label(
            frame_esquerda,
            text="NOME:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_nome.grid(row=2, column=0, padx=10, pady=5, sticky=W)
        e_nome = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=30)
        e_nome.grid(row=2, column=1, padx=10, pady=5, sticky=W)
        e_nome.bind("<KeyRelease>", converter_maiusculas)

        # Título "Telefone" zzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzz
        l_tel = Label(
            frame_esquerda,
            text="TEL:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_tel.grid(row=3, column=0, padx=10, pady=5, sticky=W)
        e_tel = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=30)
        e_tel.grid(row=3, column=1, padx=10, pady=5, sticky=W)

        # Título "CPF/CNPJ" zzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzz
        l_cpf_cnpj = Label(
            frame_esquerda,
            text="CPF/CNPJ:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_cpf_cnpj.grid(row=4, column=0, padx=10, pady=5, sticky=W)
        e_cpf_cnpj = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=30)
        e_cpf_cnpj.grid(row=4, column=1, padx=10, pady=5, sticky=W)

        # xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
        # Separador para "Dados da Entrega"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_dados_entrega = Label(
            frame_esquerda,
            text="DADOS DA ENTREGA",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 12, "bold"),
        )
        l_dados_entrega.grid(row=5, column=0, columnspan=2, padx=10, pady=10, sticky=W)

        # Título "Status"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_status = Label(
            frame_esquerda,
            text="STATUS:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_status.grid(row=6, column=0, padx=10, pady=5, sticky=W)
        e_status = ttk.Combobox(
            frame_esquerda,
            values=lista_status,
            background=cod_cor15,
            foreground="black",
            width=30,
        )
        e_status.grid(row=6, column=1, padx=10, pady=5, sticky=W)

        # Habilitar o preenchimento automático conforme você digita
        e_status["state"] = "normal"
        e_status.set("")  # Limpar o campo inicialmente
        # Definir a lista de opções inicialmente
        e_status["values"] = lista_status

        # Função para mostrar a lista suspensa sem travar o campo de entrada
        def show_dropdown(event):
            typed_text = e_status.get()

            # Apenas exibir a lista de sugestões, sem modificar o que está digitado
            if len(typed_text) >= 3:
                # Filtrar vendedores que contenham a sequência digitada
                filtered_values = [
                    v for v in lista_status if typed_text.lower() in v.lower()
                ]
                if filtered_values:
                    e_status["values"] = filtered_values  # Exibir sugestões filtradas
                    e_status.event_generate("<Down>")  # Mostrar a lista suspensa
                else:
                    e_status["values"] = (
                        []
                    )  # Se não houver correspondências, esvaziar a lista
            else:
                e_status["values"] = (
                    []
                )  # Limpar a lista suspensa se houver menos de 3 caracteres

        # Validação ao sair do campo, apenas aceita valores da lista
        def validate_choice(event=None):
            typed_text = e_status.get()
            if typed_text not in lista_status:
                pass

        # Função para autocompletar com o "Tab", sem interferir na digitação normal
        def auto_complete(event):
            typed_text = e_status.get()
            if typed_text == "":
                e_status["values"] = lista_status
            else:
                # Filtra a lista de acordo com o que foi digitado
                suggestions = [
                    v for v in lista_status if typed_text.lower() in v.lower()
                ]
                if suggestions:
                    e_status.set(suggestions[0])  # Completa com a primeira sugestão
                    e_status.icursor(
                        len(typed_text)
                    )  # Mantém o cursor após o texto digitado
                    pass

        # Bind para mostrar sugestões conforme o usuário digita
        e_status.bind("<FocusOut>", validate_choice)
        # Bind para mostrar sugestões conforme o usuário digita
        e_status.bind("<KeyRelease>", show_dropdown)
        e_status.bind("<KeyRelease>", converter_maiusculas)
        # Evento para completar com "Tab"
        e_status.bind("<Tab>", auto_complete, validate_choice)
        # Evento para completar com "Tab"
        e_status.bind("<Enter>", auto_complete, validate_choice)

        # Definir a lista de opções inicialmente e limpar o campo
        e_status["state"] = "normal"
        e_status.set("")

        # Título "Endereço"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_endereco = Label(
            frame_esquerda,
            text="ENDEREÇO:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_endereco.grid(row=7, column=0, padx=10, pady=5, sticky=W)
        e_endereco = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=30)
        e_endereco.grid(row=7, column=1, padx=10, pady=5, sticky=W)
        e_endereco.bind("<KeyRelease>", converter_maiusculas)

        # Título "Cidade"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_cidade = Label(
            frame_esquerda,
            text="CIDADE:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_cidade.grid(row=8, column=0, padx=10, pady=5, sticky=W)
        e_cidade = ttk.Combobox(
            frame_esquerda,
            values=lista_cidades,
            background=cod_cor15,
            foreground="black",
            width=30,
        )
        e_cidade.grid(row=8, column=1, padx=10, pady=5, sticky=W)

        # Habilitar o preenchimento automático conforme você digita
        e_cidade["state"] = "normal"
        e_cidade.set("")  # Limpar o campo inicialmente

        # Função para mostrar a lista suspensa sem travar o campo de entrada
        def show_dropdown(event):
            typed_text = e_cidade.get()

            # Apenas exibir a lista de sugestões, sem modificar o que está digitado
            if len(typed_text) >= 3:
                # Filtrar cidades que contenham a sequência digitada
                filtered_values = [
                    v for v in lista_cidades if typed_text.lower() in v.lower()
                ]
                if filtered_values:
                    e_cidade["values"] = filtered_values  # Exibir sugestões filtradas
                    e_cidade.event_generate("<Down>")  # Mostrar a lista suspensa
                else:
                    e_cidade["values"] = (
                        lista_cidades  # Mostrar todas as opções se não houver correspondências
                    )
            else:
                e_cidade["values"] = (
                    lista_cidades  # Mostrar todas as opções se houver menos de 3 caracteres
                )

        # Validação ao sair do campo, apenas aceita valores da lista
        def validate_choice(event=None):
            typed_text = e_cidade.get()
            if typed_text not in lista_cidades and typed_text != "":
                e_cidade.set("")  # Limpa o campo se a escolha for inválida

        # Função para completar com "Tab", sem interferir na digitação normal
        def auto_complete(event):
            typed_text = e_cidade.get()
            if typed_text == "":  # Se o campo estiver vazio, mostrar todas as opções
                e_cidade["values"] = lista_cidades
            else:
                # Filtra a lista de acordo com o que foi digitado
                suggestions = [
                    v for v in lista_cidades if typed_text.lower() in v.lower()
                ]
                if suggestions:
                    e_cidade.set(suggestions[0])  # Completa com a primeira sugestão
                    e_cidade.icursor(
                        len(typed_text)
                    )  # Mantém o cursor após o texto digitado

        # Bind para mostrar sugestões conforme o usuário digita
        e_cidade.bind("<KeyRelease>", show_dropdown)

        # Bind para validar a entrada ao perder o foco
        e_cidade.bind("<FocusOut>", validate_choice)

        # Evento para completar com "Tab"
        e_cidade.bind("<Tab>", auto_complete)

        # Definir a lista de opções inicialmente e limpar o campo
        e_cidade["state"] = "normal"
        e_cidade.set("")

        # Título "Localização"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_localizacao = Label(
            frame_esquerda,
            text="LOCALIZAÇÃO:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_localizacao.grid(row=9, column=0, padx=10, pady=5, sticky=W)
        e_localizacao = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=30)
        e_localizacao.grid(row=9, column=1, padx=10, pady=5, sticky=W)


        # Título "Ref"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_ref = Label(
            frame_esquerda,
            text="REF:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_ref.grid(row=10, column=0, padx=10, pady=5, sticky=W)
        e_ref = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=30)
        e_ref.grid(row=10, column=1, padx=10, pady=5, sticky=W)
        e_ref.bind("<KeyRelease>", converter_maiusculas)

        # Título "Carro/Ano"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_carro_ano = Label(
            frame_esquerda,
            text="CARRO/ANO:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_carro_ano.grid(row=11, column=0, padx=10, pady=5, sticky=W)
        e_carro_ano = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=30)
        e_carro_ano.grid(row=11, column=1, padx=10, pady=5, sticky=W)
        e_carro_ano.bind("<KeyRelease>", converter_maiusculas)

        # Título "Bateria"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_bateria = Label(
            frame_esquerda,
            text="BATERIA:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_bateria.grid(row=12, column=0, padx=10, pady=5, sticky=W)
        e_bateria = ttk.Combobox(
            frame_esquerda,
            values=lista_baterias,
            background=cod_cor15,
            foreground="black",
            width=10,
        )
        e_bateria.grid(row=12, column=1, padx=10, pady=5, sticky=W)

        # Habilitar o preenchimento automático conforme você digita
        e_bateria["state"] = "normal"
        e_bateria.set("")  # Limpar o campo inicialmente
        # Definir a lista de opções inicialmente
        e_bateria["values"] = lista_baterias

        # Função para mudar o foco para o próximo widget
        def focus_next_widget(event):
            # Fecha a lista suspensa antes de mover o foco
            e_bateria.event_generate("<Escape>")  # Fechar a lista suspensa
            event.widget.tk_focusNext().focus()
            return "break"  # Interrompe o comportamento padrão do Enter para evitar outras ações

        # Função para mostrar a lista suspensa sem travar o campo de entrada
        def show_dropdown(event):
            typed_text = e_bateria.get()

            # Apenas exibir a lista de sugestões, sem modificar o que está digitado
            if len(typed_text) >= 3:
                # Filtrar vendedores que contenham a sequência digitada
                filtered_values = [
                    v for v in lista_baterias if typed_text.lower() in v.lower()
                ]
                if filtered_values:
                    e_bateria["values"] = filtered_values  # Exibir sugestões filtradas
                    e_bateria.event_generate("<Down>")  # Mostrar a lista suspensa
                else:
                    e_bateria["values"] = (
                        []
                    )  # Se não houver correspondências, esvaziar a lista
            else:
                e_bateria["values"] = (
                    []
                )  # Limpar a lista suspensa se houver menos de 3 caracteres

        # Validação ao sair do campo, apenas aceita valores da lista
        def validate_choice(event=None):
            typed_text = e_bateria.get()
            if typed_text not in lista_baterias:
                pass

        # Função para autocompletar com o "Tab", sem interferir na digitação normal
        def auto_complete(event):
            typed_text = e_bateria.get()
            if typed_text == "":
                e_bateria["values"] = lista_baterias
            else:
                # Filtra a lista de acordo com o que foi digitado
                suggestions = [
                    v for v in lista_baterias if typed_text.lower() in v.lower()
                ]
                if suggestions:
                    e_bateria.set(suggestions[0])  # Completa com a primeira sugestão
                    e_bateria.icursor(
                        len(typed_text)
                    )  # Mantém o cursor após o texto digitado
                    pass

        # Bind para detectar Enter e mover para o próximo campo
        e_bateria.bind("<Return>", focus_next_widget)

        # Mostrar sugestões conforme o usuário digita
        e_bateria.bind("<KeyRelease>", show_dropdown)

        # Validar a escolha ao sair do campo
        e_bateria.bind("<FocusOut>", validate_choice)

        # Simular o comportamento de Tab ao pressionar Enter
        e_bateria.bind("<Tab>", auto_complete)

        # Inicializar com estado 'normal' e sem valor
        e_bateria["state"] = "normal"
        e_bateria.set("")

        # Título "VALOR/FORM"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_valor_form = Label(
            frame_esquerda,
            text="VALOR/FORM:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_valor_form.grid(row=13, column=0, padx=10, pady=5, sticky=W)
        e_valor_form = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=10)
        e_valor_form.grid(row=13, column=1, padx=10, pady=5, sticky=W)

        # Título "FORMA PGT"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ

        e_forma_pgt = ttk.Combobox(frame_esquerda, values=lista_forma_pgt, width=16)
        e_forma_pgt.grid(row=13, column=1, padx=93, pady=5, sticky=W)
        # Habilitar o preenchimento automático conforme você digita
        e_forma_pgt["state"] = "normal"
        # Limpar o campo inicialmente
        # Definir a lista de opções inicialmente

        # Função para mudar o foco para o próximo widget
        def focus_next_widget(event):
            # Fecha a lista suspensa antes de mover o foco
            e_forma_pgt.event_generate("<Escape>")  # Fechar a lista suspensa
            event.widget.tk_focusNext().focus()
            return "break"  # Interrompe o comportamento padrão do Enter para evitar outras ações

        # Função para mostrar a lista suspensa sem travar o campo de entrada
        def show_dropdown(event):
            typed_text = e_forma_pgt.get()

            # Apenas exibir a lista de sugestões, sem modificar o que está digitado
            if len(typed_text) >= 3:
                # Filtrar vendedores que contenham a sequência digitada
                filtered_values = [
                    v for v in lista_forma_pgt if typed_text.lower() in v.lower()
                ]
                if filtered_values:
                    e_forma_pgt["values"] = (
                        filtered_values  # Exibir sugestões filtradas
                    )
                    e_forma_pgt.event_generate("<Down>")  # Mostrar a lista suspensa
                else:
                    e_forma_pgt["values"] = (
                        []
                    )  # Se não houver correspondências, esvaziar a lista
            else:
                e_forma_pgt["values"] = (
                    []
                )  # Limpar a lista suspensa se houver menos de 3 caracteres

        # Validação ao sair do campo, apenas aceita valores da lista
        def validate_choice(event=None):
            typed_text = e_forma_pgt.get()
            if typed_text not in lista_forma_pgt:
                pass

        # Função para autocompletar com o "Tab", sem interferir na digitação normal
        def auto_complete(event):
            typed_text = e_forma_pgt.get()
            if typed_text == "":
                e_forma_pgt["values"] = lista_forma_pgt
            else:
                # Filtra a lista de acordo com o que foi digitado
                suggestions = [
                    v for v in lista_forma_pgt if typed_text.lower() in v.lower()
                ]
                if suggestions:
                    e_forma_pgt.set(suggestions[0])  # Completa com a primeira sugestão
                    e_forma_pgt.icursor(
                        len(typed_text)
                    )  # Mantém o cursor após o texto digitado
                    pass

        # Bind para detectar Enter e mover para o próximo campo
        e_forma_pgt.bind("<Return>", focus_next_widget)

        # Mostrar sugestões conforme o usuário digita
        e_forma_pgt.bind("<KeyRelease>", show_dropdown)

        # Validar a escolha ao sair do campo
        e_forma_pgt.bind("<FocusOut>", validate_choice)

        # Simular o comportamento de Tab ao pressionar Enter
        e_forma_pgt.bind("<Tab>", auto_complete)

        # Inicializar com estado 'normal' e sem valor
        e_forma_pgt["state"] = "normal"
        e_forma_pgt.set("")

        # Título "BASE DE TROCA"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_troca = Label(
            frame_esquerda,
            text="BASE DE TROCA:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_troca.grid(row=14, column=0, padx=10, pady=5, sticky=W)
        e_troca = ttk.Combobox(
            frame_esquerda,
            values=lista_troca,
            background=cod_cor15,
            foreground="black",
            width=30,
        )
        e_troca.grid(row=14, column=1, padx=10, pady=5, sticky=W)

        # Habilitar o preenchimento automático conforme você digita
        e_troca["state"] = "normal"
        e_troca.set("")  # Limpar o campo inicialmente
        # Definir a lista de opções inicialmente
        e_troca["values"] = lista_troca

        # Função para mostrar a lista suspensa sem travar o campo de entrada
        def show_dropdown(event):
            typed_text = e_troca.get()

            # Apenas exibir a lista de sugestões, sem modificar o que está digitado
            if len(typed_text) >= 3:
                # Filtrar vendedores que contenham a sequência digitada
                filtered_values = [
                    v for v in lista_troca if typed_text.lower() in v.lower()
                ]
                if filtered_values:
                    e_troca["values"] = filtered_values  # Exibir sugestões filtradas
                    e_troca.event_generate("<Down>")  # Mostrar a lista suspensa
                else:
                    e_troca["values"] = (
                        []
                    )  # Se não houver correspondências, esvaziar a lista
            else:
                e_troca["values"] = (
                    []
                )  # Limpar a lista suspensa se houver menos de 3 caracteres

        # Validação ao sair do campo, apenas aceita valores da lista
        def validate_choice(event=None):
            typed_text = e_troca.get()
            if typed_text not in lista_troca:
                pass

        # Função para autocompletar com o "Tab", sem interferir na digitação normal
        def auto_complete(event):
            typed_text = e_troca.get()
            if typed_text == "":
                e_troca["values"] = lista_troca
            else:
                # Filtra a lista de acordo com o que foi digitado
                suggestions = [
                    v for v in lista_troca if typed_text.lower() in v.lower()
                ]
                if suggestions:
                    e_troca.set(suggestions[0])  # Completa com a primeira sugestão
                    e_troca.icursor(
                        len(typed_text)
                    )  # Mantém o cursor após o texto digitado
                    pass

        # Bind para mostrar sugestões conforme o usuário digita
        e_troca.bind("<FocusOut>", validate_choice)
        # Bind para mostrar sugestões conforme o usuário digita
        e_troca.bind("<KeyRelease>", show_dropdown)

        # Evento para completar com "Tab"
        e_troca.bind("<Tab>", auto_complete, validate_choice)
        # Evento para completar com "Tab"
        e_troca.bind("<Enter>", auto_complete, validate_choice)

        # Definir a lista de opções inicialmente e limpar o campo
        e_troca["state"] = "normal"
        e_troca.set("")

        # Título "PRAZO"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_prazo = Label(
            frame_esquerda,
            text="PRAZO:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_prazo.grid(row=15, column=0, padx=10, pady=5, sticky=W)
        e_prazo = ttk.Combobox(
            frame_esquerda,
            values=lista_prazo,
            background=cod_cor15,
            foreground="black",
            width=30,
        )
        e_prazo.grid(row=15, column=1, padx=10, pady=5, sticky=W)

        # Habilitar o preenchimento automático conforme você digita
        e_prazo["state"] = "normal"
        e_prazo.set("")  # Limpar o campo inicialmente
        # Definir a lista de opções inicialmente
        e_prazo["values"] = lista_status

        # Função para mostrar a lista suspensa sem travar o campo de entrada
        def show_dropdown(event):
            typed_text = e_prazo.get()

            # Apenas exibir a lista de sugestões, sem modificar o que está digitado
            if len(typed_text) >= 3:
                # Filtrar vendedores que contenham a sequência digitada
                filtered_values = [
                    v for v in lista_prazo if typed_text.lower() in v.lower()
                ]
                if filtered_values:
                    e_prazo["values"] = filtered_values  # Exibir sugestões filtradas
                    e_prazo.event_generate("<Down>")  # Mostrar a lista suspensa
                else:
                    e_prazo["values"] = (
                        []
                    )  # Se não houver correspondências, esvaziar a lista
            else:
                e_prazo["values"] = (
                    []
                )  # Limpar a lista suspensa se houver menos de 3 caracteres

        # Validação ao sair do campo, apenas aceita valores da lista
        def validate_choice(event=None):
            typed_text = e_prazo.get()
            if typed_text not in lista_prazo:
                pass

        # Função para autocompletar com o "Tab", sem interferir na digitação normal
        def auto_complete(event):
            typed_text = e_prazo.get()
            if typed_text == "":
                e_prazo["values"] = lista_prazo
            else:
                # Filtra a lista de acordo com o que foi digitado
                suggestions = [
                    v for v in lista_prazo if typed_text.lower() in v.lower()
                ]
                if suggestions:
                    e_prazo.set(suggestions[0])  # Completa com a primeira sugestão
                    e_prazo.icursor(
                        len(typed_text)
                    )  # Mantém o cursor após o texto digitado
                    pass

        # Bind para mostrar sugestões conforme o usuário digita
        e_prazo.bind("<FocusOut>", validate_choice)
        # Bind para mostrar sugestões conforme o usuário digita
        e_prazo.bind("<KeyRelease>", show_dropdown)

        # Evento para completar com "Tab"
        e_prazo.bind("<Tab>", auto_complete, validate_choice)
        # Evento para completar com "Tab"
        e_prazo.bind("<Enter>", auto_complete, validate_choice)

        # Definir a lista de opções inicialmente e limpar o campo
        e_prazo["state"] = "normal"
        e_prazo.set("")

        # Título "MIDIA"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_midia = Label(
            frame_esquerda,
            text="MIDIA:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_midia.grid(row=16, column=0, padx=10, pady=5, sticky=W)
        e_midia = ttk.Combobox(
            frame_esquerda,
            values=lista_midias,
            background=cod_cor15,
            foreground="black",
            width=30,
        )
        e_midia.grid(row=16, column=1, padx=10, pady=5, sticky=W)

        # Habilitar o preenchimento automático conforme você digita
        e_midia["state"] = "normal"
        e_midia.set("")  # Limpar o campo inicialmente
        # Definir a lista de opções inicialmente
        e_midia["values"] = lista_midias

        # Função para mostrar a lista suspensa sem travar o campo de entrada
        def show_dropdown(event):
            typed_text = e_midia.get()

            # Apenas exibir a lista de sugestões, sem modificar o que está digitado
            if len(typed_text) >= 3:
                # Filtrar vendedores que contenham a sequência digitada
                filtered_values = [
                    v for v in lista_midias if typed_text.lower() in v.lower()
                ]
                if filtered_values:
                    e_midia["values"] = filtered_values  # Exibir sugestões filtradas
                    e_midia.event_generate("<Down>")  # Mostrar a lista suspensa
                else:
                    e_midia["values"] = (
                        []
                    )  # Se não houver correspondências, esvaziar a lista
            else:
                e_midia["values"] = (
                    []
                )  # Limpar a lista suspensa se houver menos de 3 caracteres

        # Validação ao sair do campo, apenas aceita valores da lista
        def validate_choice(event=None):
            typed_text = e_midia.get()
            if typed_text not in lista_midias:
                pass

        # Função para autocompletar com o "Tab", sem interferir na digitação normal
        def auto_complete(event):
            typed_text = e_midia.get()
            if typed_text == "":
                e_midia["values"] = lista_midias
            else:
                # Filtra a lista de acordo com o que foi digitado
                suggestions = [
                    v for v in lista_midias if typed_text.lower() in v.lower()
                ]
                if suggestions:
                    e_midia.set(suggestions[0])  # Completa com a primeira sugestão
                    e_midia.icursor(
                        len(typed_text)
                    )  # Mantém o cursor após o texto digitado
                    pass

        # Bind para mostrar sugestões conforme o usuário digita
        e_midia.bind("<FocusOut>", validate_choice)
        # Bind para mostrar sugestões conforme o usuário digita
        e_midia.bind("<KeyRelease>", show_dropdown)

        # Evento para completar com "Tab"
        e_midia.bind("<Tab>", auto_complete, validate_choice)
        # Evento para completar com "Tab"
        e_midia.bind("<Enter>", auto_complete, validate_choice)

        # Definir a lista de opções inicialmente e limpar o campo
        e_midia["state"] = "normal"
        e_midia.set("")

        # Título "RETORNO"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_retorno = Label(
            frame_esquerda,
            text="RETORNO:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_retorno.grid(row=17, column=0, padx=10, pady=5, sticky=W)
        e_retorno = ttk.Combobox(
            frame_esquerda,
            values=lista_retorno,
            background=cod_cor15,
            foreground="black",
            width=30,
        )
        e_retorno.grid(row=17, column=1, padx=10, pady=5, sticky=W)

        # Habilitar o preenchimento automático conforme você digita
        e_retorno["state"] = "normal"
        e_retorno.set("")  # Limpar o campo inicialmente
        # Definir a lista de opções inicialmente
        e_retorno["values"] = lista_status

        # Função para mostrar a lista suspensa sem travar o campo de entrada
        def show_dropdown(event):
            typed_text = e_retorno.get()

            # Apenas exibir a lista de sugestões, sem modificar o que está digitado
            if len(typed_text) >= 3:
                # Filtrar vendedores que contenham a sequência digitada
                filtered_values = [
                    v for v in lista_retorno if typed_text.lower() in v.lower()
                ]
                if filtered_values:
                    e_retorno["values"] = filtered_values  # Exibir sugestões filtradas
                    e_retorno.event_generate("<Down>")  # Mostrar a lista suspensa
                else:
                    e_retorno["values"] = (
                        []
                    )  # Se não houver correspondências, esvaziar a lista
            else:
                e_retorno["values"] = (
                    []
                )  # Limpar a lista suspensa se houver menos de 3 caracteres

        # Validação ao sair do campo, apenas aceita valores da lista
        def validate_choice(event=None):
            typed_text = e_retorno.get()
            if typed_text not in lista_retorno:
                pass

        # Função para autocompletar com o "Tab", sem interferir na digitação normal
        def auto_complete(event):
            typed_text = e_retorno.get()
            if typed_text == "":
                e_retorno["values"] = lista_retorno
            else:
                # Filtra a lista de acordo com o que foi digitado
                suggestions = [
                    v for v in lista_retorno if typed_text.lower() in v.lower()
                ]
                if suggestions:
                    e_retorno.set(suggestions[0])  # Completa com a primeira sugestão
                    e_retorno.icursor(
                        len(typed_text)
                    )  # Mantém o cursor após o texto digitado
                    pass

        # Bind para mostrar sugestões conforme o usuário digita
        e_retorno.bind("<FocusOut>", validate_choice)
        # Bind para mostrar sugestões conforme o usuário digita
        e_retorno.bind("<KeyRelease>", show_dropdown)

        # Evento para completar com "Tab"
        e_retorno.bind("<Tab>", auto_complete, validate_choice)
        # Evento para completar com "Tab"
        e_retorno.bind("<Enter>", auto_complete, validate_choice)

        # Definir a lista de opções inicialmente e limpar o campo
        e_retorno["state"] = "normal"
        e_retorno.set("")

        # Título "TABELA"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_tabela = Label(
            frame_esquerda,
            text="TABELA:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_tabela.grid(row=18, column=0, padx=10, pady=5, sticky=W)
        e_tabela = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=10)
        e_tabela.grid(row=18, column=1, padx=10, pady=5, sticky=W)
        e_tabela.bind("<KeyRelease>", converter_maiusculas)
        ################################################################
        # CANAL ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ

        # Criando a variável para armazenar a escolha do canal
        var_canal = StringVar()
        var_canal.set("")  # Inicialmente sem seleção

        # Título "CANAL"
        l_canal = Label(
            frame_esquerda,
            text="CANAL:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_canal.grid(row=19, column=0, padx=10, pady=5, sticky=W)

        # Criando o Combobox para seleção do canal
        e_canal = ttk.Combobox(
            frame_esquerda,
            values=lista_canais,
            background=cod_cor15,
            foreground="black",
            width=30,
        )
        e_canal.grid(row=19, column=1, padx=10, pady=5, sticky=W)

        # Habilitar o preenchimento automático conforme o usuário digita
        e_canal["state"] = "normal"
        e_canal.set("")  # Limpar o campo inicialmente

        # Função para mostrar a lista suspensa sem travar o campo de entrada
        def show_dropdown_canal(event):
            typed_text = e_canal.get()

            # Exibir a lista de sugestões conforme o usuário digita, sem alterar o que foi digitado
            if len(typed_text) >= 3:
                # Filtrar canais que contenham a sequência digitada
                filtered_values = [
                    v for v in lista_canais if typed_text.lower() in v.lower()
                ]
                if filtered_values:
                    e_canal["values"] = filtered_values  # Exibir sugestões filtradas
                    e_canal.event_generate("<Down>")  # Mostrar a lista suspensa
                else:
                    e_canal["values"] = (
                        []
                    )  # Esvaziar a lista se não houver correspondências
            else:
                e_canal["values"] = (
                    []
                )  # Limpar a lista suspensa se houver menos de 3 caracteres

        # Validação ao sair do campo, apenas aceita valores da lista
        def validate_choice_canal(event=None):
            typed_text = e_canal.get()
            if typed_text not in lista_canais:
                e_canal.set(
                    ""
                )  # Limpa o campo se o valor digitado não estiver na lista

        # Função para autocompletar com o "Tab", sem interferir na digitação normal
        def auto_complete_canal(event):
            typed_text = e_canal.get()
            if typed_text == "":
                e_canal["values"] = lista_canais
            else:
                # Filtra a lista de acordo com o que foi digitado
                suggestions = [
                    v for v in lista_canais if typed_text.lower() in v.lower()
                ]
                if suggestions:
                    e_canal.set(suggestions[0])  # Completa com a primeira sugestão
                    e_canal.icursor(
                        len(typed_text)
                    )  # Mantém o cursor após o texto digitado

        # Bind para mostrar sugestões conforme o usuário digita
        e_canal.bind("<FocusOut>", validate_choice_canal)
        e_canal.bind("<KeyRelease>", show_dropdown_canal)

        # Evento para completar com "Tab"
        e_canal.bind("<Tab>", auto_complete_canal)
        e_canal.bind("<Enter>", auto_complete_canal)

        # Definir a lista de opções inicialmente e limpar o campo
        e_canal["state"] = "normal"
        e_canal.set("")  # Limpar o campo inicialmente
        # Título "QUEM IRÁ RECEBER"ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        l_quem_ira_receber = Label(
            frame_esquerda,
            text="QUEM IRÁ RECEBER:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_quem_ira_receber.grid(row=20, column=0, padx=10, pady=5, sticky=W)
        e_quem_ira_receber = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=30)
        e_quem_ira_receber.grid(row=20, column=1, padx=10, pady=5, sticky=W)
        e_quem_ira_receber.bind("<KeyRelease>", converter_maiusculas)

        # Título "OBS"
        l_obs = Label(
            frame_esquerda,
            text="OBS:",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 10, "bold"),
        )
        l_obs.grid(row=21, column=0, padx=10, pady=5, sticky=W)
        e_obs = Entry(frame_esquerda, bg=cod_cor15, fg="black", width=30)
        e_obs.grid(row=21, column=1, padx=10, pady=5, sticky=W)
        e_obs.bind("<KeyRelease>", converter_maiusculas)

        # XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX

        ##CRIANDO BOTOES-----------------------------------------------------
        # Aqui para colocar as ações no botão é necessário colocar a função command = comando especifico
        # Botão "Lançar na Planilha"
        b_lancar = Button(
            frame_esquerda,
            command=inserir,
            text="LANÇAR NA PLANILHA",
            bg=cod_cor01,
            fg=cod_cor15,
            font=("Verdana", 12, "bold"),
        )
        b_lancar.grid(row=22, column=0, columnspan=2, padx=150, pady=10, sticky=W)

        # BOTAO ATUALIZAR ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        # Colocando Logo
        img_atualizar = Image.open(resource_path("imgs/atualizada.png"))
        img_atualizar = img_atualizar.resize((20, 20))  # tamanho imagem
        img_atualizar = ImageTk.PhotoImage(img_atualizar)

        # botao atualizar
        b_atualizar = Button(
            frame_esquerda,
            command=atualizar,
            width=15,
            text="atualizar ".upper(),
            compound=LEFT,
            anchor=NW,
            font=("Ivy 10 bold"),
            overrelief=RIDGE,
            bg=cod_cor22,
            fg=cod_cor28,
        )
        b_atualizar.grid(row=3, column=1, padx=255, pady=5)

        # BOTAO DELETAR ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        # Colocando Logo
        img_deletar = Image.open(resource_path("imgs/delete.png"))
        img_deletar = img_deletar.resize((20, 20))  # tamanho imagem
        img_deletar = ImageTk.PhotoImage(img_deletar)

        # botao DELETAR
        b_deletar = Button(
            frame_esquerda,
            command=deletar,
            width=15,
            text="deletar".upper(),
            compound=LEFT,
            anchor=NW,
            font=("Ivy 10 bold"),
            overrelief=RIDGE,
            bg=cod_cor22,
            fg=cod_cor28,
        )
        b_deletar.grid(row=4, column=1, padx=25, pady=5)

        # BOTAO DOWNLOAD ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        # Defina o caminho do banco de dados e o nome da tabela
        db_path = r"\\192.168.16.4\central de vendas\cadapp\atendimento_wpp_teste1.db"
        table_name = "lancamentos_whatsapp_teste1"
        # Colocando Logo
        img_dowload = Image.open(os.path.join(base_dir_atual, "imgs", "baixar.png"))
        img_dowload = img_dowload.resize((20, 20))  # tamanho imagem
        img_dowload = ImageTk.PhotoImage(img_dowload)

        # botao download
        b_dowload = Button(
            frame_esquerda,
            command=lambda: download_table_with_date_filter(db_path, table_name),
            width=15,
            text="  download ".upper(),
            compound=LEFT,
            anchor=NW,
            font=("Ivy 10 bold"),
            overrelief=RIDGE,
            bg=cod_cor22,
            fg=cod_cor28,
        )
        b_dowload.grid(row=5, column=1, padx=25, pady=5)

        # BOTAO TEXTO WPP ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        # titulo botao gerar texto wpp
        # Colocando Logo
        img_txt_wpp = Image.open(resource_path("imgs/whatsapp.png"))
        img_txt_wpp = img_txt_wpp.resize((20, 20))  # tamanho imagem
        img_txt_wpp = ImageTk.PhotoImage(img_txt_wpp)

        # botao ver texto wpp
        b_txt_wpp = Button(
            frame_esquerda,
            command=gerar_mensagem_formatada,
            width=15,
            text="ver texto wpp".upper(),
            compound=LEFT,
            anchor=NW,
            font=("Ivy 10 bold"),
            overrelief=RIDGE,
            bg=cod_cor22,
            fg=cod_cor28,
        )
        b_txt_wpp.grid(row=6, column=1, padx=25, pady=5)
        # botao ver texto wpp
        b_limpar_campos = Button(
            frame_esquerda,
            command=limpar_campos,
            width=15,
            text="limpar campos".upper(),
            compound=LEFT,
            anchor=NW,
            font=("Ivy 10 bold"),
            overrelief=RIDGE,
            bg=cod_cor22,
            fg=cod_cor28,
        )
        b_limpar_campos.grid(row=7, column=1, padx=25, pady=5)

        def open_new():
            # Configuração de cores
            VERDE = '#398564'  # Verde sugerido
            BRANCO = '#f9f7f3'  # Branco sugerido

            def processar_email():
                """Processa o e-mail e formata no estilo desejado."""
                conteudo_email = entrada_email_texto.get("1.0", tk.END)
                linhas = conteudo_email.splitlines()  # Dividir o texto em linhas

                # Dicionário para armazenar os resultados
                resultado_dict = {
                    "vendedor": "",
                    "data_horario": "",
                    "protocolo": "",
                    "cliente": "",
                    "telefone01": "",
                    "cpf_cnpj": "",
                    "endereco": "",
                    "cidade": "",
                    "localizacao": "",
                    "carro_ano": "",
                    "bateria": "",
                    "valor": "",
                    "pagamento": "",
                    "base_troca": "",
                    "prazo": "",
                }

                def buscar_termo(linhas, termo):
                    for linha in linhas:
                        if termo in linha:
                            # Extração usando regex para obter o valor após o termo
                            resultado = re.search(rf"{re.escape(termo)}\s*:\s*(.*)", linha)
                            if resultado:
                                return resultado.group(1).strip()  # Retorna apenas o conteúdo após o termo
                    return ""  # Retornar vazio se o termo não for encontrado na linha

                # Buscar cada termo individualmente, linha por linha
                resultado_dict["vendedor"] = buscar_termo(linhas, "NOME DO AGENTE")
                resultado_dict["data_horario"] = buscar_termo(linhas, "DATA/HORÁRIO")
                resultado_dict["protocolo"] = buscar_termo(linhas, "PROTOCOLO")
                resultado_dict["cliente"] = buscar_termo(linhas, "NOME DO CONTATO")
                resultado_dict["telefone01"] = buscar_termo(linhas, "NÚMERO ATENDIDO")
                resultado_dict["cpf_cnpj"] = buscar_termo(linhas, "CPF/CNPJ:")
                resultado_dict["endereco"] = buscar_termo(linhas, "ENDEREÇO:")
                resultado_dict["cidade"] = buscar_termo(linhas, "CIDADE:")
                resultado_dict["localizacao"] = buscar_termo(linhas, "OBSERVAÇÕES")
                resultado_dict["carro_ano"] = buscar_termo(linhas, "CARRO/ANO:")
                resultado_dict["bateria"] = buscar_termo(linhas, "BATERIA:")
                resultado_dict["valor"] = buscar_termo(linhas, "VALOR:")
                resultado_dict["pagamento"] = buscar_termo(linhas, "F. PAGAMENTO:")
                resultado_dict["base_troca"] = buscar_termo(linhas, "BASE DE TROCA:")
                resultado_dict["prazo"] = buscar_termo(linhas, "PRAZO:")

                # Ajuste de formatação de data/hora para remover segundos
                data_horario_formatado = resultado_dict["data_horario"][:-3] if resultado_dict["data_horario"] else ""

                # Formatando os dados para o formato desejado
                resultado = f"""
*VENDEDOR:* {resultado_dict['vendedor']}
*DATA/HORÁRIO:* {data_horario_formatado}

*COD:* {resultado_dict['protocolo']}
*CLIENTE:* {resultado_dict['cliente']}
*TELEFONE 01:* {resultado_dict['telefone01']}
*CPF/CNPJ:* {resultado_dict['cpf_cnpj']}

*ENDEREÇO:* {resultado_dict['endereco']}
*CIDADE:* {resultado_dict['cidade']}
*LOCALIZAÇÃO:* {resultado_dict['localizacao']}
*REF:* 

*CARRO/ANO:* {resultado_dict['carro_ano']}
*BATERIA:* {resultado_dict['bateria']}
*VALOR:* R$ {resultado_dict['valor'].replace('.', ',') if resultado_dict['valor'] else ''}
*F. PAGAMENTO:* {resultado_dict['pagamento']}
*BASE DE TROCA:* {resultado_dict['base_troca']}
*PRAZO:* {resultado_dict['prazo']}

*QUEM IRÁ RECEBER:* {resultado_dict['cliente']}
                """
                
                # Exibe o resultado formatado na caixa de saída
                saida_texto.delete("1.0", tk.END)
                saida_texto.insert(tk.END, resultado.strip())

                # Copia o resultado para a área de transferência
                pyperclip.copy(resultado.strip())
                messagebox.showinfo("Copiado", "Resultado formatado copiado para a área de transferência!")

            def limpar_email():
                """Limpa o campo de entrada do e-mail e a saída."""
                entrada_email_texto.delete("1.0", tk.END)
                saida_texto.delete("1.0", tk.END)

            # ------------------------
            # INTERFACE GRÁFICA (TKINTER)
            # ------------------------

            # Configuração da interface tkinter
            janela_opn = tk.Toplevel()  # Usar Toplevel para abrir essa janela como um pop-up
            janela_opn.title("Conversor texto Open New")
            janela_opn.geometry("750x600")
            janela_opn.configure(bg="#398564")  # Cor de fundo principal verde escuro

            # Estrutura principal para organizar a interface
            frame_principal = tk.Frame(janela_opn, bg="#398564")
            frame_principal.pack(fill=tk.BOTH, expand=True)

            # Lado esquerdo para entrada e saída de texto
            frame_texto = tk.Frame(frame_principal, bg="#398564")
            frame_texto.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            # Caixa de entrada para o e-mail
            tk.Label(frame_texto, text="Cole o conteúdo do Open New abaixo e processe:", fg="#f9f7f3", bg="#398564", font=("Arial", 12, "bold")).pack(anchor="w", padx=10)
            entrada_email_texto = scrolledtext.ScrolledText(frame_texto, wrap=tk.WORD, width=60, height=15, font=("Arial", 12))
            entrada_email_texto.pack(pady=10, padx=10, anchor="w")

            # Caixa de saída para o resultado
            saida_texto = scrolledtext.ScrolledText(frame_texto, wrap=tk.WORD, width=60, height=15, font=("Arial", 12))
            saida_texto.pack(pady=10, padx=10, anchor="w")

            # Lado direito para botões
            frame_botoes = tk.Frame(frame_principal, bg="#398564")
            frame_botoes.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)

            # Botão para processar o e-mail
            botao_processar = tk.Button(frame_botoes, text="Processar E-mail", command=processar_email, bg="#f9f7f3", fg="#398564", font=("Arial", 10, "bold"))
            botao_processar.pack(pady=10, fill=tk.X)

            # Botão para limpar o campo de entrada e saída
            botao_limpar = tk.Button(frame_botoes, text="Limpar", command=limpar_email, bg="#f9f7f3", fg="#398564", font=("Arial", 10, "bold"))
            botao_limpar.pack(pady=10, fill=tk.X)

            # Botão para fechar a janela
            botao_fechar = tk.Button(frame_botoes, text="Fechar Open New", command=janela_opn.destroy, bg="#f9f7f3", fg="#398564", font=("Arial", 10, "bold"))
            botao_fechar.pack(pady=10, fill=tk.X)

            # Inicia a interface
            janela_opn.mainloop()



                # botao ver texto wpp
        b_texto_opennew = Button(
            frame_esquerda,
            command=open_new,
            width=15,
            text="Texto Open New".upper(),
            compound=LEFT,
            anchor=NW,
            font=("Ivy 10 bold"),
            overrelief=RIDGE,
            bg=cod_cor22,
            fg=cod_cor28,
        )
        b_texto_opennew.grid(row=8, column=1, padx=25, pady=5)

        

        # Adiciona o botão de fechar ao final da interface
        btn_fechar = tk.Button(
            frame_direita_cima, text="Fechar Aplicativo", command=fechar_app
        )
        btn_fechar.grid(row=0, column=1, padx=25, pady=5)
        # CRIANDO CARTÕES/KPIS

        ##CAIXAS DE TEXTO / LABELS VALORES E QTD TOTAL
        # Data de hoje
        hoje = datetime.date.today()
        data_formatada_hoje = hoje.strftime("%d/%m/%Y")


        # Função para calcular a soma das compras do dia atual diretamente do banco de dados
        def calcular_soma_dia_atual():
            con = sqlite3.connect(db_path)
            cur = con.cursor()
            
            # Query para somar o preço das compras onde a data é a de hoje
            query_soma = """
                SELECT SUM(preco) FROM lancamentos_whatsapp_teste1
                WHERE strftime('%Y-%m-%d', data) = date('now')
            """
            
            cur.execute(query_soma)
            soma_total_dia = cur.fetchone()[0]
            con.close()

            # Se a soma for None (nenhum valor encontrado), definimos como 0.0
            if soma_total_dia is None:
                soma_total_dia = 0.0

            return soma_total_dia

        
        # Função para contar a quantidade de NFs do dia atual diretamente do banco de dados
        def contar_qtd_nf_dia_atual():
            con = sqlite3.connect(db_path)
            cur = con.cursor()
            
            # Query para contar a quantidade de NFs onde a data é a de hoje
            query_qtd = """
                SELECT COUNT(*) FROM lancamentos_whatsapp_teste1
                WHERE strftime('%Y-%m-%d', data) = date('now')
            """
            
            cur.execute(query_qtd)
            qtd_nf_dia = cur.fetchone()[0]
            con.close()

            return qtd_nf_dia

        # Função para calcular a soma das compras do mês atual (26 do mês anterior até 25 do mês atual)
        def calcular_soma_mes_atual():
            hoje = datetime.date.today()

            # Determinar os dias de início e fim do "mês atual"
            dia_atual = hoje.day
            mes_atual = hoje.month
            ano_atual = hoje.year

            if dia_atual >= 26:
                # Estamos no "mês atual" (26 do mês anterior até 25 do mês atual)
                inicio_mes_atual = datetime.date(ano_atual, mes_atual, 26)
                fim_mes_atual = (inicio_mes_atual + datetime.timedelta(days=30)).replace(day=25)
            else:
                # Estamos no período que conta para o mês anterior
                fim_mes_atual = datetime.date(ano_atual, mes_atual, 25)
                mes_anterior = mes_atual - 1 if mes_atual > 1 else 12
                ano_anterior = ano_atual if mes_atual > 1 else ano_atual - 1
                inicio_mes_atual = datetime.date(ano_anterior, mes_anterior, 26)

            # Conectar ao banco e fazer a consulta para somar os valores dentro desse intervalo de datas
            con = sqlite3.connect(db_path)
            cur = con.cursor()

            # Query para somar o preço das compras no intervalo do mês atual
            query_soma_mes = """
                SELECT SUM(preco) FROM lancamentos_whatsapp_teste1
                WHERE date(data) BETWEEN ? AND ?
            """
            cur.execute(query_soma_mes, (inicio_mes_atual, fim_mes_atual))
            soma_total_mes = cur.fetchone()[0]
            con.close()

            # Se a soma for None (nenhum valor encontrado), definimos como 0.0
            if soma_total_mes is None:
                soma_total_mes = 0.0

            return soma_total_mes


        # Caixa TOTAL DO DIA
        l_total_dia_titulo = Label(
            frame_direita_cima,
            text=f"Valores de hoje: {data_formatada_hoje}",
            height=1,
            anchor=CENTER,
            font=("Ivy 13 bold"),
            bg=cod_cor18,
            fg=cod_cor16,
        )
        l_total_dia_titulo.grid(
            row=0, column=0, padx=10, pady=(10, 5), sticky="ew"
        )  # Usando grid para centralizar

        # Exibindo o valor calculado
        l_total_dia_caixa = Label(
            frame_direita_cima,
            text="",
            width=16,
            height=2,
            anchor=CENTER,
            font=("Ivy 17 bold"),
            bg=cod_cor18,
            fg=cod_cor16,
        )
        l_total_dia_caixa.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

        # Caixa QUANTIDADE DO DIA
        l_qtd_dia_titulo = Label(
            frame_direita_cima,
            text=f"QTD. de NFs: {data_formatada_hoje}",
            height=1,
            anchor=CENTER,
            font=("Ivy 13 bold"),
            bg=cod_cor18,
            fg=cod_cor16,
        )
        l_qtd_dia_titulo.grid(row=2, column=0, padx=10, pady=(20, 5), sticky="ew")

        l_qtd_dia_caixa = Label(
            frame_direita_cima,
            text="",
            width=16,
            height=2,
            anchor=CENTER,
            font=("Ivy 17 bold"),
            bg=cod_cor18,
            fg=cod_cor16,
        )
        l_qtd_dia_caixa.grid(row=3, column=0, padx=10, pady=5, sticky="ew")

        # Caixa SOMA DO MÊS ATUAL
        l_total_mes_titulo = Label(
            frame_direita_cima,
            text=f"Valores do mês atual: {data_formatada_hoje}",
            height=1,
            anchor=CENTER,
            font=("Ivy 13 bold"),
            bg=cod_cor18,
            fg=cod_cor16,
        )
        l_total_mes_titulo.grid(row=4, column=0, padx=10, pady=(20, 5), sticky="ew")

        l_total_mes_caixa = Label(
            frame_direita_cima,
            text="",
            width=16,
            height=2,
            anchor=CENTER,
            font=("Ivy 17 bold"),
            bg=cod_cor18,
            fg=cod_cor16,
        )
        l_total_mes_caixa.grid(row=5, column=0, padx=10, pady=5, sticky="ew")


        # Configurar a distribuição de espaço no frame para expandir horizontalmente
        frame_direita_cima.grid_columnconfigure(0, weight=1)



        ##CRIANDO A TABELA PARA VISUALIZAR ###################################################################################################################



        # FILTROS #####################################################################################################################################

        import datetime
        from tkinter import ttk

        from tkcalendar import Calendar, DateEntry

        # Função para validar e converter a data de string para formato datetime ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        def converter_data(data_str):
            try:
                return datetime.datetime.strptime(data_str, "%d/%m/%Y").date()
            except ValueError:
                return None

        # Função para aplicar o filtro de atendente ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        def aplicar_filtro():
            filtro_atendente = (
                combo_filtro_atendente.get()
            )  # Obtém o atendente selecionado no combobox
            mostrar(filtro_atendente)

        # Função para remover os filtros ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        def remover_filtros():
            combo_filtro_atendente.set("")  # Limpa o filtro de atendente
            mostrar()  # Mostra todos os registros novamente

        # Função para preencher a lista de atendentes no combobox (incluindo lista estática) ZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZZ
        def preencher_lista_atendentes():
            con = sqlite3.connect(db_path)
            cur = con.cursor()

            # Pegar todos os atendentes únicos do banco de dados
            cur.execute("SELECT DISTINCT atendente FROM lancamentos_whatsapp_teste1")
            atendentes_db = [row[0] for row in cur.fetchall()]
            con.close()

            # Combina a lista estática com a lista do banco (sem duplicados)
            atendentes_completos = list(set(lista_vendedores + atendentes_db))

            return atendentes_completos

        def get_date_range_mostrar():
            def grab_dates():
                """Pega as datas e fecha a janela se estiverem válidas."""
                start_date = start_entry.get()
                end_date = end_entry.get()

                # Verificar o formato correto (dd/mm/yyyy)
                if len(start_date) != 10 or len(end_date) != 10:
                    messagebox.showwarning(
                        "Erro", "As datas devem estar no formato dd/mm/yyyy."
                    )
                    return None, None

                # Converter as datas para o formato SQL (yyyy-mm-dd)
                try:
                    start_date = pd.to_datetime(start_date, dayfirst=True).strftime(
                        "%Y-%m-%d"
                    )
                    end_date = pd.to_datetime(end_date, dayfirst=True).strftime(
                        "%Y-%m-%d"
                    )
                except ValueError:
                    messagebox.showerror(
                        "Erro", "Datas inválidas. Por favor, insira datas válidas."
                    )
                    return None, None

                root.destroy()  # Fechar a janela
                return start_date, end_date

            def close_window():
                """Fecha a janela e retorna None para cancelar a operação."""
                root.destroy()  # Fecha a janela completamente
                return None, None

            def set_today():
                """Define a data de hoje para ambos os campos."""
                today = pd.to_datetime("today").strftime("%d/%m/%Y")
                start_entry.delete(0, tk.END)
                start_entry.insert(0, today)
                end_entry.delete(0, tk.END)
                end_entry.insert(0, today)

            def set_current_month():
                """Define o mês atual (início e fim do mês)."""
                first_day = pd.to_datetime("today").replace(day=1).strftime("%d/%m/%Y")
                last_day = (
                    pd.to_datetime("today").replace(day=1)
                    + pd.DateOffset(months=1)
                    - pd.DateOffset(days=1)
                )
                last_day = last_day.strftime("%d/%m/%Y")
                start_entry.delete(0, tk.END)
                start_entry.insert(0, first_day)
                end_entry.delete(0, tk.END)
                end_entry.insert(0, last_day)

            def set_current_year():
                """Define o ano atual (início e fim do ano)."""
                first_day_year = (
                    pd.to_datetime("today").replace(month=1, day=1).strftime("%d/%m/%Y")
                )
                last_day_year = (
                    pd.to_datetime("today")
                    .replace(month=12, day=31)
                    .strftime("%d/%m/%Y")
                )
                start_entry.delete(0, tk.END)
                start_entry.insert(0, first_day_year)
                end_entry.delete(0, tk.END)
                end_entry.insert(0, last_day_year)

            def set_all_until_today():
                """Define todas as datas desde o início até hoje."""
                first_date_db = "01/01/1900"
                today = pd.to_datetime("today").strftime("%d/%m/%Y")
                start_entry.delete(0, tk.END)
                start_entry.insert(0, first_date_db)
                end_entry.delete(0, tk.END)
                end_entry.insert(0, today)

            # Criar janela para o intervalo de datas
            root = tk.Toplevel()
            root.title("Selecione o intervalo de datas")

            # Centralizando a janela
            largura_janela = 400
            altura_janela = 250
            largura_tela = root.winfo_screenwidth()
            altura_tela = root.winfo_screenheight()
            pos_x = (largura_tela // 2) - (largura_janela // 2)
            pos_y = (altura_tela // 2) - (altura_janela // 2)
            root.geometry(f"{largura_janela}x{altura_janela}+{pos_x}+{pos_y}")

            # Layout para data inicial
            Label(root, text="Data Inicial (dd/mm/yyyy):").grid(
                row=0, column=0, padx=10, pady=5
            )
            start_entry = Entry(root, width=12)
            start_entry.grid(row=0, column=1)

            # Layout para data final
            Label(root, text="Data Final (dd/mm/yyyy):").grid(
                row=1, column=0, padx=10, pady=5
            )
            end_entry = Entry(root, width=12)
            end_entry.grid(row=1, column=1)

            # Botões de intervalo rápido
            Button(root, text="Hoje", command=set_today).grid(
                row=0, column=2, padx=5, pady=5
            )
            Button(root, text="Mês Atual", command=set_current_month).grid(
                row=1, column=2, padx=5, pady=5
            )
            Button(root, text="Ano Atual", command=set_current_year).grid(
                row=2, column=2, padx=5, pady=5
            )
            Button(root, text="Tudo até Hoje", command=set_all_until_today).grid(
                row=3, column=2, padx=5, pady=5
            )

            # Botão para confirmar
            confirm_button = Button(root, text="Confirmar", command=lambda: root.quit())
            confirm_button.grid(row=4, column=0, pady=10)

            # Botão para voltar
            back_button = Button(root, text="Voltar", command=close_window)
            back_button.grid(row=4, column=1, pady=10)

            root.mainloop()

            return grab_dates()  # Retornar as datas selecionadas ou None

        # Função para aplicar o filtro de data usando a função get_date_range_mostrar
        def aplicar_filtro_data():
            start_date, end_date = get_date_range_mostrar()  # Obter intervalo de datas
            if start_date and end_date:
                mostrar(
                    start_date=start_date, end_date=end_date
                )  # Aplicar o filtro de data

        # LABELS E BOTÕES PARA FILTROS #######################################################################################################################

        # Criação do Combobox e Botão para aplicar o filtro na parte superior do frame_direita_baixo
        l_filtro_atendente = tk.Label(
            frame_direita_baixo,
            text="Atendente:",
            bg=cod_cor15,
            fg="black",
            font=("Ivy 10"),
        )
        l_filtro_atendente.place(x=10, y=5)

        # Preencher o combobox com a lista de atendentes
        combo_filtro_atendente = ttk.Combobox(
            frame_direita_baixo, values=preencher_lista_atendentes(), width=15
        )
        combo_filtro_atendente.place(x=90, y=5)

        # Botão para aplicar o filtro
        b_aplicar_filtro = tk.Button(
            frame_direita_baixo,
            text="Aplicar Filtro",
            command=aplicar_filtro,
            width=15,
            font=("Ivy 10 bold"),
            bg=cod_cor18,
            fg=cod_cor16,
        )
        b_aplicar_filtro.place(x=250, y=5)


        # Botão para remover os filtros
        b_remover_filtro = tk.Button(
            frame_direita_baixo,
            text="Remover Filtros",
            command=remover_filtros,
            width=15,
            font=("Ivy 10 bold"),
            bg=cod_cor18,
            fg=cod_cor16,
        )
        b_remover_filtro.place(x=400, y=5)

        # Label e botão para selecionar o intervalo de datas
        l_data_range = Label(
            frame_direita_baixo,
            text="Intervalo de Datas:",
            bg=cod_cor15,
            fg="black",
            font=("Ivy 10"),
        )
        l_data_range.place(x=10, y=40)
        # Botão para abrir o calendário de seleção de datas
        b_selecionar_data = Button(
            frame_direita_baixo,
            text="Selecionar Datas",
            command=lambda: aplicar_filtro_data(),
            width=20,
            font=("Ivy 10 bold"),
            bg=cod_cor18,
            fg=cod_cor16,
        )
        b_selecionar_data.place(x=150, y=40)


        # Função para exibir os dados no Treeview
        def mostrar(filtro_atendente=None, start_date=None, end_date=None):
            global tree

            # Cabeçalhos da tabela
            tabela_head = [
                "ID",
                "Data",
                "Hora",
                "Atendente",
                "COD",
                "Cliente",
                "Telefone",
                "CPF/CNPJ",
                "Status",
                "Endereço",
                "Cidade",
                "Localização",
                "REF",
                "Carro",
                "Produto",
                "Preço",
                "Base Troca",
                "Prazo",
                "Mídia Social",
                "Retorno",
                "Tabela",
                "Canal",
                "Quem Irá Receber",
                "Observações",
            ]

            # Criando o Treeview que ficará na metade inferior do frame (permitindo rolagem para mais dados)
            tree = ttk.Treeview(
                frame_direita_baixo,
                selectmode="extended",
                columns=tabela_head,
                show="headings",
                height=10,
            )

            # Barra de rolagem vertical e horizontal
            vsb = ttk.Scrollbar(
                frame_direita_baixo, orient="vertical", command=tree.yview
            )
            hsb = ttk.Scrollbar(
                frame_direita_baixo, orient="horizontal", command=tree.xview
            )

            # Configurações da barra de rolagem no Treeview
            tree.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)

            # Posicionando o Treeview e as barras de rolagem
            tree.place(
                relx=0, rely=0.5, relwidth=1, relheight=0.45
            )  # Colocando o Treeview na metade inferior
            vsb.place(
                relx=0.98, rely=0.5, relheight=0.45
            )  # Barra de rolagem vertical na lateral direita
            hsb.place(
                relx=0, rely=0.95, relwidth=0.98
            )  # Barra de rolagem horizontal na parte inferior

            # Configurações das colunas
            hd = ["center"] * len(tabela_head)
            h = [
                50,
                100,
                80,
                120,
                80,
                150,
                120,
                120,
                120,
                200,
                120,
                150,
                80,
                120,
                100,
                80,
                100,
                80,
                120,
                100,
                100,
                100,
                150,
                250,
            ]

            for idx, col in enumerate(tabela_head):
                tree.heading(col, text=col, anchor=tk.CENTER)
                tree.column(col, width=h[idx], anchor=hd[idx], stretch=False)

            # Conectar ao banco de dados
            con = sqlite3.connect(db_path)
            cur = con.cursor()

                        # Montar a query com base nos filtros
            query = """
                SELECT id, strftime('%d/%m/%Y', data), hora, atendente, cod, cliente, telefone, cpf_cnpj, status, endereco, cidade, localizacao, ref, carro, produto, preco, base_troca, prazo, midia_social, retorno, tabela, canal, quem_ira_receber, observacoes
                FROM lancamentos_whatsapp_teste1
                WHERE 1=1
               
            """

            params = []

            if filtro_atendente:
                query += " AND atendente = ?"
                params.append(filtro_atendente)

            if start_date and end_date:
                query += " AND date(data) BETWEEN ? AND ?"
                params.append(start_date)
                params.append(end_date)
            # Ordenar pelo ID em ordem decrescente
            query += " ORDER BY strftime('%Y/%m/%d', data) DESC, time(hora) DESC"

            cur.execute(query, params)
            lista_itens = cur.fetchall()
            con.close()

            # Limpar a tabela antes de inserir os itens
            for i in tree.get_children():
                tree.delete(i)

            # Inserir os itens no Treeview
            for item in lista_itens:
                tree.insert("", "end", values=item)

            # Atualiza a interface
            tree.update_idletasks()

            # Atualizar KPIs (valores e quantidade de NFs do dia)
            soma_total_hoje = calcular_soma_dia_atual()
            qtd_nf_hoje = contar_qtd_nf_dia_atual()
            soma_total_mes = calcular_soma_mes_atual()  # Soma do mês atual

            # Atualizando os labels de soma e quantidade de NFs com formatação de milhares e duas casas decimais
            l_total_dia_caixa.config(text=f"R$ {locale.format_string('%.2f', soma_total_hoje, grouping=True)}")
            l_qtd_dia_caixa.config(text=f"{qtd_nf_hoje} NFs")
            l_total_mes_caixa.config(text=f"R$ {locale.format_string('%.2f', soma_total_mes, grouping=True)}")


        # Exibir a tabela pela primeira vez (sem filtro)
        mostrar()
        
        tabela_head = [
                "ID",
                "Data",
                "Hora",
                "Atendente",
                "COD",
                "Cliente",
                "Telefone",
                "CPF/CNPJ",
                "Status",
                "Endereço",
                "Cidade",
                "Localização",
                "REF",
                "Carro",
                "Produto",
                "Preço",
                "Base Troca",
                "Prazo",
                "Mídia Social",
                "Retorno",
                "Tabela",
                "Canal",
                "Quem Irá Receber",
                "Observações",
            ]

        # Função para redefinir colunas para o tamanho padrão
        def redefinir_colunas():
            # Larguras padrão das colunas
            larguras_iniciais = [
                50, 100, 80, 120, 80, 150, 120, 120, 120, 200, 120, 150, 80, 120, 100,
                80, 100, 80, 120, 100, 100, 100, 150, 250
            ]
            for idx, col in enumerate(tabela_head):
                tree.column(col, width=larguras_iniciais[idx])  # Redefinir larguras para o padrão

        # Função para atualizar a tabela sem alterar os dados
        def atualizar_tabela():
            mostrar()  # Simplesmente chama a função mostrar() para fazer o "refresh" da tabela

        # Criando o botão de "Redefinir Colunas" no frame direito, logo acima da tabela, no lado esquerdo

        botao_redefinir = Button(
            frame_direita_baixo,
            text="Redefinir Col.",
            command=redefinir_colunas,
            width=10,
            font=("Ivy 10 bold"),
            bg='grey',
            fg="#f9f7f3",
        )
        botao_redefinir.grid(row=4, column=5, padx=50, pady=70, sticky="ew")
   

        # Criando o botão de "Atualizar Tabela" no frame direito, logo acima da tabela, no lado esquerdo, abaixo do outro botão

        botao_refresh = Button(
            frame_direita_baixo,
            text="Atualizar Vis.",
            command=atualizar_tabela,
            width=10,
            font=("Ivy 10 bold"),
            bg='grey',
            fg="#f9f7f3",
        )
        botao_refresh.grid(row=4, column=2, padx=10, pady=70, sticky="ew")
   

        janela.mainloop()
    except ImportError as e:
        messagebox.showerror("Erro de Importação", f"Erro ao importar módulos: {e}")
    except sqlite3.Error as e:
        messagebox.showerror(
            "Erro no Banco de Dados", f"Erro ao acessar o banco de dados: {e}"
        )
    except Exception as e:
        messagebox.showerror("Erro Desconhecido", f"Ocorreu um erro inesperado: {e}")


####### BOTÕES ##############################################################
import tkinter as tk

# Login
login_button = ttk.Button(frame_direita, text="Login", width=30, command=loggin)
login_button.place(x=135, y=225)

# registrar
registrar_button = ttk.Button(
    frame_direita, text="Registrar", width=20, command=register
)
registrar_button.place(x=170, y=260)


janela_loggin.mainloop()
